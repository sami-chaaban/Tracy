from ._shared import *
from .. import __version__

class TrajectoryCanvas(QWidget):
    def __init__(self, parent=None, kymo_canvas=None, movie_canvas=None, intensity_canvas=None, navigator=None):
        super().__init__(parent)
        self.setFocusPolicy(Qt.StrongFocus)
        self.intensity_canvas = intensity_canvas
        self.kymoCanvas = kymo_canvas
        self.movieCanvas = movie_canvas
        self.navigator = navigator
        self.trajectories = []  # List of trajectory dicts.
        self._trajectory_counter = 1

        # Table widget for displaying trajectory summary information.
        self.table_widget = QTableWidget()
        # 1) Enable alternating row colors
        self.table_widget.setAlternatingRowColors(True)

        # 2) Tweak header appearance
        header = self.table_widget.horizontalHeader()
        header.setContextMenuPolicy(Qt.CustomContextMenu)
        header.customContextMenuRequested.connect(self._on_header_context_menu)
        header.setSectionsClickable(False)
        header.setHighlightSections(False)
        header.setDefaultAlignment(Qt.AlignCenter)

        # 4) hide the grid or only show horizontal lines:
        self.table_widget.setShowGrid(False)
        self.table_widget.setAlternatingRowColors(True)
        self.table_widget.setSelectionBehavior(QTableWidget.SelectRows)
        self.table_widget.setColumnCount(13)
        self.table_widget.setMinimumHeight(50)
        self._headers = [
            "", "Channel", "Frame A", "Frame B",
            "Start X,Y", "End X,Y",
            "Distance μm", "Time s", "Net Speed μm/s",
            "Total", "Valid %",
            "Med. Intensity",
        ]
        self.table_widget.setColumnCount(len(self._headers))
        self.table_widget.setHorizontalHeaderLabels(self._headers)

        # 2) map header text → column index
        self._col_index = { hdr: idx
                            for idx, hdr in enumerate(self._headers) }

        # 3) aliases
        self._aliases = {
            "channel":      "Channel",
            "startframe":   "Frame A",
            "endframe":     "Frame B",
            "startcoord":   "Start X,Y",
            "endcoord":     "End X,Y",
            "distance":     "Distance μm",
            "time":         "Time s",
            "netspeed":     "Net Speed μm/s",
            "total":        "Total",
            "valid":        "Valid %",
            "medintensity":"Med. Intensity",
        }
        self.table_widget.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.table_widget.verticalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.table_widget.verticalHeader().setVisible(False)
        self.table_widget.setSelectionBehavior(QTableWidget.SelectRows)
        self.table_widget.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table_widget.itemSelectionChanged.connect(self.on_trajectory_selected_by_table)
        # For the first column, adjust its width to fit its contents:
        self.table_widget.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)

        # For the rest of the columns (indexes 1 to n), set them as interactive:
        for i in range(1, self.table_widget.columnCount()): 
            if i in [1,2,3,7,9,10]:
                self.table_widget.horizontalHeader().setSectionResizeMode(i, QHeaderView.Interactive)
                self.table_widget.setColumnWidth(i, 70)
            elif i in [4,5,6]:
                self.table_widget.horizontalHeader().setSectionResizeMode(i, QHeaderView.Interactive)
                self.table_widget.setColumnWidth(i, 100)              
            elif i in [8,11]:
                self.table_widget.horizontalHeader().setSectionResizeMode(i, QHeaderView.Interactive)
                self.table_widget.setColumnWidth(i, 120)              
            else:
                self.table_widget.horizontalHeader().setSectionResizeMode(i, QHeaderView.Interactive)
                self.table_widget.setColumnWidth(i, 130)

        # turn on mouse‑tracking so we get hover events
        # self.table_widget.setMouseTracking(True)
        # # cellEntered gives us (row, col) whenever the cursor enters a cell
        # self.table_widget.cellEntered.connect(self.on_table_cell_hovered)
        # install an event filter on the viewport so we can catch Leave

        self.table_widget.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_widget.customContextMenuRequested.connect(self.open_context_menu)

        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        self.top_divider = QFrame()
        self.top_divider.setObjectName("TrajectoryTopLine")
        self.top_divider.setFixedHeight(1)
        self.top_divider.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.top_divider.setStyleSheet(
            "QFrame#TrajectoryTopLine { background-color: #d0d0d0; }"
        )
        layout.addWidget(self.top_divider)
        layout.addWidget(self.table_widget)
        self.setLayout(layout)

        self.custom_columns = []
        self._column_types = {}
        self._custom_load_map = {}

        self.current_index = None

    def makeCenteredItem(self, text):
        item = QTableWidgetItem(text)
        item.setTextAlignment(Qt.AlignCenter)
        return item

    def writeToTable(self, row, key, text):
        """
        key can be the exact header text or an alias.
        """
        # Special case for column 0 if needed.
        if key == "num":
            col = 0
        else:
            # normalize lookup
            key_lower = key.lower()

            # 1) try alias
            header = self._aliases.get(key_lower)
            # 2) fallback to exact header match
            if header is None:
                if key in self._col_index:
                    header = key
                else:
                    raise KeyError(f"No column or alias found for '{key}'")

            col = self._col_index[header]

        item = self.makeCenteredItem(text)
        if col == 0:
            font = item.font()
            font.setBold(True)
            item.setFont(font)

        # set the item
        self.table_widget.setItem(row, col, item)

    # def change_trajectory_id(self, rows):
    #     """
    #     rows: list of table‐row indices to rename.
    #     For simplicity we only allow renaming *one* at a time:
    #     """
    #     if len(rows) != 1:
    #         QMessageBox.information(
    #             self, "Bulk rename not supported",
    #             "Please select exactly one trajectory to rename."
    #         )
    #         return

    #     row = rows[0]
    #     old_id_item = self.table_widget.item(row, 0)
    #     old_id = int(old_id_item.text())

    #     # Ask for a new ID
    #     new_id, ok = QInputDialog.getInt(
    #         self, "Change Trajectory ID",
    #         f"Enter new ID for trajectory {old_id}:",
    #         value=old_id,
    #         min=0,
    #         max=1_000_000
    #     )
    #     if not ok or new_id == old_id:
    #         return

    #     # Check for collisions
    #     existing_ids = {traj["trajectory_number"] for traj in self.trajectories}
    #     if new_id in existing_ids:
    #         QMessageBox.warning(
    #             self, "ID Already Exists",
    #             f"A trajectory with ID {new_id} already exists.\n"
    #             "Please choose a different ID."
    #         )
    #         return

    #     # Update the in‐memory data
    #     for traj in self.trajectories:
    #         if traj["trajectory_number"] == old_id:
    #             traj["trajectory_number"] = new_id
    #             break

    #     # Refresh the whole table so rows sort by the new ID order
    #     self.refresh_trajectory_table()

    # def refresh_trajectory_table(self):
    #     #NEED TO WRITE

    def save_selected_trajectories(self):
        """
        Gather selected rows and call the unified export helper.
        """
        rows = [idx.row() for idx in self.table_widget.selectionModel().selectedRows()]
        if not rows:
            return

        # delegate to the single-export method
        self.save_trajectories(rows)

    def save_trajectories(self, rows: Optional[List[int]] = None):
        """
        Export either all trajectories (rows=None) or only those at the given row-indices.
        """
        traj_list = self.trajectories if rows is None else [self.trajectories[r] for r in rows]

        # Handle no trajectories case
        if not traj_list:
            nav_rois = set(self.navigator.rois.keys())
            if nav_rois:
                save_empty = (
                    QMessageBox.question(
                        self,
                        "",
                        "There are no trajectories. Save the empty kymographs in the Per-kymograph sheet?",
                        QMessageBox.Yes | QMessageBox.No,
                        QMessageBox.Yes
                    ) == QMessageBox.Yes
                )
                if not save_empty:
                    return
            else:
                QMessageBox.warning(self, "", "There are no trajectories to save.")
                return
        else:
            save_empty = False

        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Save Trajectories",
            "",
            "Excel Files (*.xlsx)"
        )
        if not filename:
            return

        try:

                # Build summary and data rows if any trajectories
            summary_rows, data_rows, per_node_rows = [], [], []

            def _infer_nodes_and_clicks(traj):
                nodes = traj.get("nodes") or []
                click_source = str(traj.get("click_source") or "").strip()
                anchors = traj.get("anchors") or []
                roi = traj.get("roi", None)

                def _dedupe_nodes(raw_nodes):
                    deduped = []
                    last_key = None
                    for item in raw_nodes:
                        if not isinstance(item, (list, tuple)) or len(item) < 3:
                            continue
                        frame, x, y = item[0], item[1], item[2]
                        try:
                            key = (int(frame), float(x), float(y))
                        except Exception:
                            continue
                        if key == last_key:
                            continue
                        deduped.append((key[0], key[1], key[2]))
                        last_key = key
                    return deduped

                if nodes:
                    if not click_source:
                        click_source = "kymograph" if (anchors and roi is not None) else "movie"
                    return _dedupe_nodes(nodes), click_source

                if anchors and roi is not None:
                    nodes = []
                    for frame_idx, xk, _yk in anchors:
                        mx, my = self.navigator.compute_roi_point(roi, xk)
                        nodes.append((int(frame_idx), float(mx), float(my)))
                    if not click_source:
                        click_source = "kymograph"
                return _dedupe_nodes(nodes), click_source

            def _anchor_index_for_frame(frame, anchor_frames):
                if not anchor_frames:
                    return ""
                if len(anchor_frames) < 2:
                    return 1
                for idx in range(len(anchor_frames) - 1):
                    if frame <= anchor_frames[idx + 1]:
                        return idx + 1
                return len(anchor_frames) - 1

            def _percent_yes(flags):
                valid = [s for s in flags if s is not None]
                if not valid:
                    return ""
                return f"{100 * sum(1 for s in valid if s == 'Yes') / len(valid):.1f}"

            def _round4(value):
                if value in ("", None):
                    return ""
                try:
                    return round(float(value), 4)
                except Exception:
                    return value

            def _autosize_worksheet(writer, sheet_name, df, max_width=60):
                try:
                    from openpyxl.utils import get_column_letter
                    from openpyxl.styles import Alignment
                except Exception:
                    return
                ws = writer.sheets.get(sheet_name)
                if ws is None or df is None:
                    return
                left_align = Alignment(horizontal="left", vertical="center")
                format_cols = {
                    idx for idx, col in enumerate(df.columns, start=1)
                    if col is not None and (
                        str(col).startswith("Percent Valid")
                        or str(col).startswith("Search Center")
                        or str(col).startswith("Original Coordinate")
                        or "Intensity" in str(col)
                        or str(col).startswith("Sigma")
                        or str(col).startswith("Peak")
                        or "Speed" in str(col)
                    )
                }
                for idx, col in enumerate(df.columns, start=1):
                    header_len = len(str(col)) if col is not None else 0
                    width = min(max_width, header_len + 2)
                    ws.column_dimensions[get_column_letter(idx)].width = width
                for row_idx, row in enumerate(ws.iter_rows(), start=1):
                    for col_idx, cell in enumerate(row, start=1):
                        cell.alignment = left_align
                        if row_idx > 1 and col_idx in format_cols:
                            cell.number_format = "0.0000"
            if traj_list:
                for traj in traj_list:
                    channel = int(traj["channel"])
                    fixed_background = traj["fixed_background"]
                    save_start_frame = int(traj["start"][0]) + 1
                    save_end_frame = int(traj["end"][0]) + 1

                    # step data
                    step_meds = traj.get("step_medians")
                    steps_available = step_meds is not None
                    step_meds_list = step_meds or []
                    num_steps = len(step_meds_list) if steps_available else 0

                    # placeholders
                    avg_step_size = ""
                    avg_step_size_bg = ""

                    if steps_available and num_steps > 1:
                        # 1) pairwise diffs between medians
                        diffs = [
                            abs(step_meds_list[i][2] - step_meds_list[i-1][2])
                            for i in range(1, num_steps)
                        ]
                        avg_step_size = round(sum(diffs) / len(diffs), 1)

                        # 2) include difference from last median → background
                        if fixed_background is not None:
                            last_med = step_meds_list[-1][2]
                            diff_bg = abs(last_med - fixed_background)
                            diffs_with_bg = diffs + [diff_bg]
                            avg_step_size_bg = round(sum(diffs_with_bg) / len(diffs_with_bg), 1)
                        else:
                            # fallback: just reuse the plain average
                            avg_step_size_bg = avg_step_size
                            
                    frames_list = traj.get("frames", [])
                    num_points = len(frames_list)
                    # Compute number of valid points: valid if intensity exists and is greater than 0.
                    intensities_list = traj.get("intensities", [])
                    valid_points = sum(1 for val in intensities_list if val is not None and val > 0)
                    percent_valid = (100 * valid_points / num_points) if num_points > 0 else 0

                    avg_vel_px_fr_txt = ""
                    avg_vel_um_s_txt = ""
                    avg_vel_um_min_txt = ""
                    if self.navigator.pixel_size is not None and self.navigator.frame_interval is not None and traj['average_velocity'] is not None:
                        avg_vel_px_fr_txt = traj["average_velocity"]
                        velocity_nm_per_ms = (traj['average_velocity'] * self.navigator.pixel_size) / self.navigator.frame_interval
                        avg_vel_um_s_txt = velocity_nm_per_ms
                        avg_vel_um_min_txt = velocity_nm_per_ms * 60.0

                    def _valid_xy(pt):
                        if not isinstance(pt, (list, tuple)) or len(pt) != 2:
                            return False
                        x, y = pt
                        if x is None or y is None:
                            return False
                        if isinstance(x, float) and math.isnan(x):
                            return False
                        if isinstance(y, float) and math.isnan(y):
                            return False
                        return True

                    search_centers = traj.get("search_centers", []) or []
                    start_center = next((pt for pt in search_centers if _valid_xy(pt)), None)
                    end_center = next((pt for pt in reversed(search_centers) if _valid_xy(pt)), None)
                    if start_center is None:
                        start_center = (traj["start"][1], traj["start"][2])
                    if end_center is None:
                        end_center = (traj["end"][1], traj["end"][2])

                    dx = end_center[0] - start_center[0]
                    dy = end_center[1] - start_center[1]
                    distance_px = np.hypot(dx, dy)
                    frames_list = traj.get("frames", [])
                    time_fr = (frames_list[-1] - frames_list[0]) if len(frames_list) > 1 else 0

                    distance_um_txt = ""
                    time_s_txt = ""
                    overall_vel_px_fr_txt = ""
                    overall_vel_um_s_txt = ""
                    overall_vel_um_min_txt = ""

                    if time_fr > 0:
                        overall_vel_px_fr = distance_px / time_fr
                        overall_vel_px_fr_txt = overall_vel_px_fr

                    if self.navigator.pixel_size is not None:
                        distance_um = distance_px * self.navigator.pixel_size / 1000
                        distance_um_txt = f"{distance_um:.2f}"

                    if self.navigator.frame_interval is not None and time_fr > 0:
                        time_s = time_fr * self.navigator.frame_interval / 1000
                        time_s_txt = f"{time_s:.2f}"
                        if self.navigator.pixel_size is not None and time_s > 0:
                            overall_vel_um_s = distance_um / time_s
                            overall_vel_um_min = overall_vel_um_s * 60.0
                            overall_vel_um_s_txt = overall_vel_um_s
                            overall_vel_um_min_txt = overall_vel_um_min

                    # 1) Serialize kymo anchors
                    anchors = traj.get("anchors", [])
                    if anchors:
                        # convert to pure-Python types
                        anchors_py = [
                            (int(frame), float(x), float(y))
                            for frame, x, y in anchors
                        ]
                        anchors_str = json.dumps(anchors_py)
                    else:
                        anchors_str = ""  # empty

                    # 2) Serialize ROI
                    roi = traj.get("roi", None)
                    if roi:
                        # convert all np.floats to floats
                        roi_clean = {
                            "type": roi["type"],
                            "x": [float(xx) for xx in roi.get("x", [])],
                            "y": [float(yy) for yy in roi.get("y", [])],
                            "points": [
                                (float(px), float(py)) for px, py in roi.get("points", [])
                            ]
                        }
                        roi_str = json.dumps(roi_clean)
                    else:
                        roi_str = "" 

                    nodes, click_source = _infer_nodes_and_clicks(traj)
                    anchor_frames = sorted({
                        int(n[0])
                        for n in nodes
                        if isinstance(n, (list, tuple)) and len(n) >= 3
                    })
                    nodes_str = ""
                    if nodes:
                        nodes_py = [
                            (float(x), float(y), int(frame) + 1)
                            for frame, x, y in nodes
                        ]
                        nodes_str = json.dumps(nodes_py)

                    row = {
                        "Movie": self.navigator.movieNameLabel.text(),
                        "Trajectory": traj.get("trajectory_number", "?"),
                        "Channel": channel,
                        "Start Frame": save_start_frame,
                        "End Frame": save_end_frame,
                        "Segments": max(len(nodes) - 1, 0),
                        "Kymo-Anchors": anchors_str,
                        "ROI":    roi_str,
                        "Clicks": click_source,
                        "Movie-Anchors": nodes_str,
                        "Total Points": num_points,
                        "Valid Points": valid_points,
                        "Percent Valid": _round4(percent_valid),
                        "Search Center X Start": "" if start_center is None else _round4(start_center[0]),
                        "Search Center Y Start": "" if start_center is None else _round4(start_center[1]),
                        "Search Center X End": "" if end_center is None else _round4(end_center[0]),
                        "Search Center Y End": "" if end_center is None else _round4(end_center[1]),
                        "Distance (μm)": distance_um_txt,
                        "Time (s)": time_s_txt,
                        "Background": fixed_background,
                        "Average Intensity": "" if traj["average"] is None else traj["average"],
                        "Median Intensity": "" if traj["median"] is None else traj["median"],
                        "Net Speed (px/frame)": _round4(overall_vel_px_fr_txt),
                        "Net Speed (μm/s)": _round4(overall_vel_um_s_txt),
                        "Net Speed (μm/min)": _round4(overall_vel_um_min_txt),
                        "Avg. Speed (px/frame)": _round4(avg_vel_px_fr_txt),
                        "Avg. Speed (μm/s)": _round4(avg_vel_um_s_txt),
                        "Avg. Speed (μm/min)": _round4(avg_vel_um_min_txt)
                    }

                    if steps_available:
                        row["Number of Steps"]    = num_steps
                        row["Average Step Size"]  = avg_step_size
                        row["Average Step Size w/Step to Background"]    = avg_step_size_bg

                    summary_rows.append(row)

                    D_COL = getattr(self.navigator, "_DIFF_D_COL", None)
                    A_COL = getattr(self.navigator, "_DIFF_A_COL", None)
                    diff_cols = {c for c in (D_COL, A_COL) if c}

                    for col in self.custom_columns:
                        col_type = self._column_types.get(col, "binary")

                        # Special cases that should be saved without a "[type]" suffix
                        if col in diff_cols:
                            header = col
                        elif col.startswith("Ch.") and col.endswith("co. %"):
                            header = col
                        else:
                            header = f"{col} [{col_type}]"

                        summary_rows[-1][header] = traj.get("custom_fields", {}).get(col, "")

                    traj_name = str(traj.get("trajectory_number", "?"))
                    coords_list = traj.get("original_coords", [])
                    centers_list = traj.get("search_centers", [])
                    spot_centers_list = traj.get("spot_centers", []) or []
                    coloc_any_list = traj.get("colocalization_any", []) or []
                    coloc_by_ch = traj.get("colocalization_by_ch", {}) or {}
                    seg_diff_by_idx = {}
                    for entry in (traj.get("segment_diffusion") or []):
                        if isinstance(entry, dict) and entry.get("segment") is not None:
                            try:
                                seg_diff_by_idx[int(entry["segment"])] = entry
                            except Exception:
                                continue
                    steps_enabled = steps_available
                    has_coloc = (
                        self.navigator.movie is not None
                        and self.navigator.movie.ndim == 4
                        and self.navigator._channel_axis is not None
                    )
                    n_chan = (
                        self.navigator.movie.shape[self.navigator._channel_axis]
                        if has_coloc else 0
                    )
                    ch_ref = channel
                    want_diffusion = bool(diff_cols) and (
                        getattr(self.navigator, "show_diffusion", False)
                        or any(c in self.custom_columns for c in diff_cols)
                    )
                    can_diffuse = (
                        want_diffusion
                        and self.navigator.pixel_size is not None
                        and self.navigator.frame_interval is not None
                    )

                    if len(anchor_frames) >= 2:
                        nodes_sorted = sorted(
                            [
                                n for n in nodes
                                if isinstance(n, (list, tuple)) and len(n) >= 3
                            ],
                            key=lambda n: n[0]
                        )

                        for node_idx in range(len(nodes_sorted) - 1):
                            start_frame, node_start_x, node_start_y = nodes_sorted[node_idx]
                            end_frame, node_end_x, node_end_y = nodes_sorted[node_idx + 1]
                            is_first_pair = (node_idx == 0)
                            seg_start = start_frame if is_first_pair else (start_frame + 1)
                            seg_indices = [
                                idx for idx, f in enumerate(frames_list)
                                if seg_start <= f <= end_frame
                            ]
                            if not seg_indices:
                                continue

                            seg_frames = [frames_list[idx] for idx in seg_indices]
                            seg_ints = [intensities_list[idx] for idx in seg_indices]
                            seg_centers = [
                                centers_list[idx] if idx < len(centers_list) else None
                                for idx in seg_indices
                            ]
                            seg_spots = [
                                spot_centers_list[idx] if idx < len(spot_centers_list) else None
                                for idx in seg_indices
                            ]
                            seg_coords = [
                                coords_list[idx] if idx < len(coords_list) else None
                                for idx in seg_indices
                            ]

                            seg_valid_points = sum(
                                1 for val in seg_ints if val is not None and val > 0
                            )
                            seg_percent_valid = (
                                (100 * seg_valid_points / len(seg_frames))
                                if seg_frames else 0
                            )

                            seg_start_center = next(
                                (pt for pt in seg_centers if _valid_xy(pt)),
                                None
                            )
                            seg_end_center = next(
                                (pt for pt in reversed(seg_centers) if _valid_xy(pt)),
                                None
                            )
                            if seg_start_center is None and seg_coords:
                                seg_start_center = seg_coords[0] if _valid_xy(seg_coords[0]) else None
                            if seg_end_center is None and seg_coords:
                                seg_end_center = seg_coords[-1] if _valid_xy(seg_coords[-1]) else None
                            if seg_start_center is None:
                                seg_start_center = (node_start_x, node_start_y)
                            if seg_end_center is None:
                                seg_end_center = (node_end_x, node_end_y)

                            seg_dx = seg_end_center[0] - seg_start_center[0]
                            seg_dy = seg_end_center[1] - seg_start_center[1]
                            seg_distance_px = np.hypot(seg_dx, seg_dy)
                            seg_time_fr = (seg_frames[-1] - seg_frames[0]) if len(seg_frames) > 1 else 0

                            seg_distance_um_txt = ""
                            seg_time_s_txt = ""
                            seg_overall_vel_px_fr_txt = ""
                            seg_overall_vel_um_s_txt = ""
                            seg_overall_vel_um_min_txt = ""

                            if seg_time_fr > 0:
                                seg_overall_vel_px_fr = seg_distance_px / seg_time_fr
                                seg_overall_vel_px_fr_txt = seg_overall_vel_px_fr

                            if self.navigator.pixel_size is not None:
                                seg_distance_um = seg_distance_px * self.navigator.pixel_size / 1000
                                seg_distance_um_txt = f"{seg_distance_um:.2f}"

                            if self.navigator.frame_interval is not None and seg_time_fr > 0:
                                seg_time_s = seg_time_fr * self.navigator.frame_interval / 1000
                                seg_time_s_txt = f"{seg_time_s:.2f}"
                                if self.navigator.pixel_size is not None and seg_time_s > 0:
                                    seg_overall_vel_um_s = seg_distance_um / seg_time_s
                                    seg_overall_vel_um_min = seg_overall_vel_um_s * 60.0
                                    seg_overall_vel_um_s_txt = seg_overall_vel_um_s
                                    seg_overall_vel_um_min_txt = seg_overall_vel_um_min

                            seg_avg_vel_px_fr_txt = ""
                            seg_avg_vel_um_s_txt = ""
                            seg_avg_vel_um_min_txt = ""
                            seg_vels = [
                                traj["velocities"][idx]
                                for idx in seg_indices
                                if idx < len(traj.get("velocities", []))
                            ]
                            valid_seg_vels = [v for v in seg_vels if v is not None]
                            if valid_seg_vels:
                                seg_avg_vel_px_fr = float(np.mean(valid_seg_vels))
                                seg_avg_vel_px_fr_txt = seg_avg_vel_px_fr
                                if self.navigator.pixel_size is not None and self.navigator.frame_interval is not None:
                                    seg_vel_nm_ms = (seg_avg_vel_px_fr * self.navigator.pixel_size) / self.navigator.frame_interval
                                    seg_avg_vel_um_s_txt = seg_vel_nm_ms
                                    seg_avg_vel_um_min_txt = seg_vel_nm_ms * 60.0

                            seg_valid_ints = [v for v in seg_ints if v is not None and v > 0]
                            seg_avg_int = float(np.mean(seg_valid_ints)) if seg_valid_ints else None
                            seg_med_int = float(np.median(seg_valid_ints)) if seg_valid_ints else None

                            seg_step_meds = []
                            seg_num_steps = 0
                            seg_avg_step_size = ""
                            seg_avg_step_size_bg = ""
                            if steps_enabled:
                                _seg_step_idxs, seg_step_meds = self.navigator.compute_steps_for_data(
                                    seg_frames, seg_ints
                                )
                                seg_num_steps = len(seg_step_meds)
                                if seg_num_steps > 1:
                                    seg_diffs = [
                                        abs(seg_step_meds[i][2] - seg_step_meds[i-1][2])
                                        for i in range(1, seg_num_steps)
                                    ]
                                    seg_avg_step_size = round(sum(seg_diffs) / len(seg_diffs), 1)
                                    if fixed_background is not None:
                                        seg_last_med = seg_step_meds[-1][2]
                                        seg_diff_bg = abs(seg_last_med - fixed_background)
                                        seg_avg_step_size_bg = round(
                                            sum(seg_diffs + [seg_diff_bg]) / (len(seg_diffs) + 1), 1
                                        )
                                    else:
                                        seg_avg_step_size_bg = seg_avg_step_size

                            seg_custom_overrides = {}
                            if want_diffusion:
                                seg_D = None
                                seg_alpha = None
                                cached = seg_diff_by_idx.get(node_idx + 1)
                                if isinstance(cached, dict):
                                    seg_D = cached.get("D")
                                    seg_alpha = cached.get("alpha")
                                if seg_D is None and seg_alpha is None:
                                    if can_diffuse:
                                        try:
                                            seg_D, seg_alpha = self.navigator.compute_diffusion_for_data(
                                                seg_frames, seg_spots
                                            )
                                        except ValueError:
                                            seg_D, seg_alpha = (None, None)

                                if D_COL:
                                    seg_custom_overrides[D_COL] = (
                                        "" if seg_D is None else f"{seg_D:.4f}"
                                    )
                                if A_COL:
                                    seg_custom_overrides[A_COL] = (
                                        "" if seg_alpha is None else f"{seg_alpha:.3f}"
                                    )

                            if has_coloc:
                                seg_any = [
                                    coloc_any_list[idx] if idx < len(coloc_any_list) else None
                                    for idx in seg_indices
                                ]
                                pct_any = _percent_yes(seg_any)
                                pct_by_ch = {}
                                for ch in range(1, n_chan + 1):
                                    flags = coloc_by_ch.get(ch, [None] * num_points)
                                    seg_flags = [
                                        flags[idx] if idx < len(flags) else None
                                        for idx in seg_indices
                                    ]
                                    pct_by_ch[ch] = _percent_yes(seg_flags)

                                for ch in range(1, n_chan + 1):
                                    col_name = f"Ch. {ch} co. %"
                                    if ch == ch_ref:
                                        val = ""
                                    elif n_chan == 2:
                                        val = pct_any
                                    else:
                                        val = pct_by_ch.get(ch, "")
                                    seg_custom_overrides[col_name] = val

                            seg_row = {
                                "Movie": self.navigator.movieNameLabel.text(),
                                "Trajectory": traj.get("trajectory_number", "?"),
                                "Segment": node_idx + 1,
                                "Channel": channel,
                                "Clicks": click_source,
                                "Segment Start X": float(node_start_x),
                                "Segment Start Y": float(node_start_y),
                                "Segment Start Frame": int(start_frame) + 1,
                                "Segment End X": float(node_end_x),
                                "Segment End Y": float(node_end_y),
                                "Segment End Frame": int(end_frame) + 1,
                                "Kymo-Anchors": anchors_str,
                                "ROI": roi_str,
                                "Total Points": len(seg_frames),
                                "Valid Points": seg_valid_points,
                                "Percent Valid": _round4(seg_percent_valid),
                                "Distance (μm)": seg_distance_um_txt,
                                "Time (s)": seg_time_s_txt,
                                "Background": fixed_background,
                                "Average Intensity": "" if seg_avg_int is None else seg_avg_int,
                                "Median Intensity": "" if seg_med_int is None else seg_med_int,
                                "Net Speed (px/frame)": _round4(seg_overall_vel_px_fr_txt),
                                "Net Speed (μm/s)": _round4(seg_overall_vel_um_s_txt),
                                "Net Speed (μm/min)": _round4(seg_overall_vel_um_min_txt),
                                "Avg. Speed (px/frame)": _round4(seg_avg_vel_px_fr_txt),
                                "Avg. Speed (μm/s)": _round4(seg_avg_vel_um_s_txt),
                                "Avg. Speed (μm/min)": _round4(seg_avg_vel_um_min_txt),
                            }

                            if steps_enabled:
                                seg_row["Number of Steps"] = seg_num_steps
                                seg_row["Average Step Size"] = seg_avg_step_size
                                seg_row["Average Step Size w/Step to Background"] = seg_avg_step_size_bg

                            for col in self.custom_columns:
                                col_type = self._column_types.get(col, "binary")

                                if col in diff_cols:
                                    header = col
                                elif col.startswith("Ch.") and col.endswith("co. %"):
                                    header = col
                                else:
                                    header = f"{col} [{col_type}]"

                                seg_row[header] = seg_custom_overrides.get(
                                    col,
                                    traj.get("custom_fields", {}).get(col, "")
                                )

                            per_node_rows.append(seg_row)

                    # Ensure we have the intensities list from above.
                    for i in range(len(frames_list)):

                        vel_nm_per_ms = ""
                        vel_um_min = ""
                        velocity = ""
                        coord_x, coord_y = "", ""
                        search_x, search_y = "", ""
                        spot_x, spot_y = "", ""
                        sigma_val = ""
                        peak_val = ""
                        background_val = ""

                        intensity_val = intensities_list[i] if (intensities_list[i] is not None and intensities_list[i] > 0) else ""
                        if i < len(coords_list):
                            coord_x, coord_y = coords_list[i]
                        else:
                            coord_x, coord_y = "", ""
                        if i < len(centers_list):
                            search_x, search_y = centers_list[i]
                        else:
                            search_x, search_y = "", ""
                        if "spot_centers" in traj and len(traj["spot_centers"]) > i and traj["spot_centers"][i] is not None:
                            spot_x, spot_y = traj["spot_centers"][i]
                        else:
                            spot_x, spot_y = "", ""
                        sigma_val = ""
                        if "sigmas" in traj and len(traj["sigmas"]) > i and traj["sigmas"][i] is not None:
                            sigma_val = traj["sigmas"][i]
                        peak_val = ""
                        if "peaks" in traj and len(traj["peaks"]) > i and traj["peaks"][i] is not None:
                            peak_val = traj["peaks"][i]
                        if "background" in traj and len(traj["background"]) > i and traj["background"][i] is not None:
                            background_val = traj["background"][i]
                        if "velocities" in traj and len(traj["velocities"]) > i and traj["velocities"][i] is not None:
                            velocity = traj["velocities"][i]
                            if self.navigator.pixel_size is not None and self.navigator.frame_interval is not None:
                                vel_nm_per_ms = (velocity * self.navigator.pixel_size) / self.navigator.frame_interval
                                vel_um_min = vel_nm_per_ms  * 60.0
                            else:
                                vel_um_min = ""
                        else:
                            velocity = ""
                            vel_nm_per_ms = ""
                            vel_um_min = ""

                        fixedstr = "No"
                        if fixed_background is not None:
                            fixedstr = "Yes"

                        f = frames_list[i]
                        node_idx = _anchor_index_for_frame(f, anchor_frames)

                        # ——— determine step number & raw step value ———
                        step_number = None
                        step_value = None
                        for sn, (start_f, end_f, median) in enumerate(step_meds_list):
                            if start_f <= f <= end_f:
                                step_number = sn
                                step_value  = median
                                break

                        # background_val is the numeric background for this frame (or "" if none)
                        if step_value is not None and background_val not in ("", None):
                            # ensure numeric subtraction
                            step_value_bg_adj = step_value - float(background_val)
                        else:
                            step_value_bg_adj = None

                        base = {
                            "Trajectory": traj_name,
                            "Segment": node_idx,
                            "Channel": channel,
                            "Clicks": click_source,
                            "Frame": frames_list[i] + 1,
                            "Original Coordinate X": _round4(coord_x),
                            "Original Coordinate Y": _round4(coord_y),
                            "Search Center X": _round4(search_x),
                            "Search Center Y": _round4(search_y),
                            "Spot Center X": spot_x,
                            "Spot Center Y": spot_y,
                            "Intensity": _round4(intensity_val),
                            "Sigma": _round4(sigma_val),
                            "Peak": _round4(peak_val),
                            "Background from trajectory": fixedstr,
                            "Background": background_val,
                            "Speed (px/frame)": _round4(velocity),
                            "Speed (μm/s)": _round4(vel_nm_per_ms),
                            "Speed (μm/min)": _round4(vel_um_min),
                        }

                        if steps_available:
                            base["Step Number"]                             = step_number
                            base["Step Intensity Value"]                    = step_value
                            base["Step Intensity Value (background-adjusted)"] = step_value_bg_adj

                        if self.navigator.movie.ndim == 4 and self.navigator._channel_axis is not None:
                            n_chan = self.navigator.movie.shape[self.navigator._channel_axis]

                            # build the coloc dict with *all* channel columns
                            coloc = {"Colocalized w/any channel": traj["colocalization_any"][i] or ""}

                            for ch in range(1, n_chan+1):
                                key = f"Colocalized w/ch{ch}"
                                # for the reference channel this will just be blank
                                flags = traj["colocalization_by_ch"].get(ch, [None]*num_points)
                                coloc[key] = flags[i] or ""

                            data_rows.append({**base, **coloc})
                        else:
                            data_rows.append(base)

            df_summary = pd.DataFrame(summary_rows)
            df_data = pd.DataFrame(data_rows)
            df_per_node = pd.DataFrame(per_node_rows)

            #per-ROI
            pixel_size_um = (
                self.navigator.pixel_size / 1000.0
                if self.navigator.pixel_size is not None else None
            )
            frame_interval_s = (
                self.navigator.frame_interval / 1000.0
                if self.navigator.frame_interval is not None else None
            )
            n_frames = self.navigator.movie.shape[0]
            total_time_s = (
                n_frames * frame_interval_s
                if frame_interval_s is not None else None
            )

            per_roi_columns = [
                "ROI",
                "Total distance (μm)",
                "Total time (s)",
                "Number of trajectories",
                "Events (/min)",
                "Events (/μm/min)",
                "Average net speed (μm/s)",
                "Average average speed (μm/s)",
                "Average run length (μm)",
                "Average run time (s)",
                "Average median intensity",
                "Average average intensity",
            ]
            all_jsons = {json.dumps(roi_dict) for roi_dict in self.navigator.rois.values()}

            def _build_per_roi_df(df_summary_subset, include_empty_rows):
                per_roi_list = []
                if not df_summary_subset.empty:
                    df_roi = df_summary_subset.copy()
                    for col in ["Distance (μm)", "Time (s)", "Net Speed (μm/s)",
                                "Avg. Speed (μm/s)", "Average Intensity", "Median Intensity"]:
                        if col in df_roi:
                            df_roi[col] = pd.to_numeric(df_roi[col], errors="coerce")

                    seen_jsons = {
                        roi for roi in df_roi["ROI"]
                        if isinstance(roi, str) and roi.strip()
                    }

                    for roi_json in seen_jsons:
                        grp = df_roi[df_roi["ROI"] == roi_json]
                        n_trajs = len(grp)
                        events_per_min = (n_trajs / (total_time_s / 60.0)) if total_time_s else float("nan")
                        total_distance_um = None
                        events_per_um_per_min = float("nan")
                        if pixel_size_um:
                            roi_dict = json.loads(roi_json)
                            xs, ys = np.array(roi_dict["x"], float), np.array(roi_dict["y"], float)
                            total_distance_um = np.sum(np.hypot(np.diff(xs), np.diff(ys))) * pixel_size_um
                            if total_time_s and total_distance_um:
                                events_per_um_per_min = events_per_min / total_distance_um

                        per_roi_list.append({
                            "ROI": roi_json,
                            "Total distance (μm)": total_distance_um,
                            "Total time (s)": total_time_s,
                            "Number of trajectories": n_trajs,
                            "Events (/min)": events_per_min,
                            "Events (/μm/min)": events_per_um_per_min,
                            "Average net speed (μm/s)": grp["Net Speed (μm/s)"].mean(),
                            "Average average speed (μm/s)": grp["Avg. Speed (μm/s)"].mean(),
                            "Average run length (μm)": grp["Distance (μm)"].mean(),
                            "Average run time (s)": grp["Time (s)"].mean(),
                            "Average median intensity": grp["Median Intensity"].mean(),
                            "Average average intensity": grp["Average Intensity"].mean(),
                        })
                else:
                    seen_jsons = set()

                empty_jsons_local = all_jsons - seen_jsons
                if empty_jsons_local and include_empty_rows:
                    for roi_json in empty_jsons_local:
                        total_distance_um = None
                        events_txt = ""
                        if pixel_size_um:
                            roi_dict = json.loads(roi_json)
                            xs, ys = np.array(roi_dict["x"], float), np.array(roi_dict["y"], float)
                            total_distance_um = np.sum(np.hypot(np.diff(xs), np.diff(ys))) * pixel_size_um
                            events_txt = "0"

                        per_roi_list.append({
                            "ROI": roi_json,
                            "Total distance (μm)": total_distance_um,
                            "Total time (s)": total_time_s,
                            "Number of trajectories": 0,
                            "Events (/min)": events_txt,
                            "Events (/μm/min)": events_txt,
                            "Average net speed (μm/s)": "",
                            "Average average speed (μm/s)": "",
                            "Average run length (μm)": "",
                            "Average run time (s)": "",
                            "Average median intensity": "",
                            "Average average intensity": "",
                        })

                return pd.DataFrame(per_roi_list, columns=per_roi_columns), seen_jsons

            # If we have trajectory-based summaries, include those
            seen_jsons = set()
            if not df_summary.empty and "ROI" in df_summary.columns:
                seen_jsons = {
                    roi for roi in df_summary["ROI"]
                    if isinstance(roi, str) and roi.strip()
                }
            empty_jsons = all_jsons - seen_jsons
            include_empty = False

            if empty_jsons:
                include_empty = (
                    QMessageBox.question(
                        self,
                        "",
                        "Include the empty kymographs in the Per-kymograph sheet?",
                        QMessageBox.Yes | QMessageBox.No,
                        QMessageBox.Yes
                    ) == QMessageBox.Yes
                )

            include_empty_rows = (save_empty or include_empty)
            df_per_roi, _ = _build_per_roi_df(df_summary, include_empty_rows)

            per_roi_by_channel = {}
            if self.navigator.movie.ndim == 4 and self.navigator._channel_axis is not None:
                n_chan = self.navigator.movie.shape[self.navigator._channel_axis]
                if "Channel" in df_summary.columns:
                    chan_series = pd.to_numeric(df_summary["Channel"], errors="coerce")
                else:
                    chan_series = None

                for ch in range(1, n_chan + 1):
                    if chan_series is None:
                        df_ch = df_summary.iloc[0:0]
                    else:
                        df_ch = df_summary[chan_series == ch]
                    per_roi_by_channel[ch], _ = _build_per_roi_df(df_ch, include_empty_rows)

            # ─────────────────────────────────────────────────────────────
            # Aggregate analysis (across ALL trajectories + ALL kymographs)
            # ─────────────────────────────────────────────────────────────

            # movie dims (use navigator's 2D frame getter if available)
            try:
                frame0 = self.navigator.get_movie_frame(0)
            except Exception:
                # fallback: best-effort from raw movie array
                if self.navigator.movie.ndim == 3:
                    frame0 = self.navigator.movie[0]
                else:
                    frame0 = self.navigator.movie[0, ...]  # may still be multi-d; best-effort

            # Ensure 2D
            if hasattr(frame0, "ndim") and frame0.ndim > 2:
                # last-resort squeeze
                frame0 = np.squeeze(frame0)

            height_px = int(frame0.shape[0])
            width_px  = int(frame0.shape[1])
            movie_dims_px_txt = f"{width_px}, {height_px}"

            px_nm = (
                float(self.navigator.pixel_size)
                if self.navigator.pixel_size is not None
                else None
            )
            ft_ms = (
                float(self.navigator.frame_interval)
                if self.navigator.frame_interval is not None
                else None
            )
            n_frames = int(self.navigator.movie.shape[0])

            total_time_s = ""
            if ft_ms is not None:
                total_time_s = (n_frames * (ft_ms / 1000.0))

            movie_dims_um_txt = ""
            if px_nm is not None:
                px_um = px_nm / 1000.0
                movie_dims_um_txt = f"{width_px * px_um:.2f}, {height_px * px_um:.2f}"

            total_kymographs = len(self.navigator.rois or {})

            # total kymograph distance (sum over ALL rois, including empty ones)
            total_kymo_distance_um = ""
            if px_nm is not None:
                pixel_size_um = px_nm / 1000.0
                total_um = 0.0
                for roi_dict in (self.navigator.rois or {}).values():
                    try:
                        xs = np.array(roi_dict.get("x", []), float)
                        ys = np.array(roi_dict.get("y", []), float)
                        if len(xs) >= 2 and len(ys) >= 2:
                            total_um += float(np.sum(np.hypot(np.diff(xs), np.diff(ys))) * pixel_size_um)
                    except Exception:
                        pass
                total_kymo_distance_um = total_um

            # how many empty kymographs? (ROIs with zero trajectories)
            all_jsons = {json.dumps(roi_dict) for roi_dict in (self.navigator.rois or {}).values()}

            if not df_summary.empty and "ROI" in df_summary.columns:
                seen_jsons = {
                    roi for roi in df_summary["ROI"]
                    if isinstance(roi, str) and roi.strip()
                }
            else:
                seen_jsons = set()

            empty_kymographs = len(all_jsons - seen_jsons)

            n_trajs = 0 if df_summary.empty else int(len(df_summary))

            events_per_min = ""
            events_per_um_per_min = ""
            if ft_ms is not None and total_time_s and total_time_s > 0:
                events_per_min = n_trajs / (total_time_s / 60.0)
                if px_nm is not None and isinstance(total_kymo_distance_um, (int, float)) and total_kymo_distance_um > 0:
                    events_per_um_per_min = events_per_min / total_kymo_distance_um

            # trajectory-wise averages (across ALL trajectories)
            avg_net_speed = ""
            avg_avg_speed = ""
            avg_run_length = ""
            avg_run_time = ""
            avg_med_int = ""
            avg_avg_int = ""

            if not df_summary.empty:
                df_all = df_summary.copy()
                for col in ["Distance (μm)", "Time (s)", "Net Speed (μm/s)",
                            "Avg. Speed (μm/s)", "Average Intensity", "Median Intensity"]:
                    if col in df_all.columns:
                        df_all[col] = pd.to_numeric(df_all[col], errors="coerce")

                if "Net Speed (μm/s)" in df_all:
                    avg_net_speed = df_all["Net Speed (μm/s)"].mean()
                if "Avg. Speed (μm/s)" in df_all:
                    avg_avg_speed = df_all["Avg. Speed (μm/s)"].mean()
                if "Distance (μm)" in df_all:
                    avg_run_length = df_all["Distance (μm)"].mean()
                if "Time (s)" in df_all:
                    avg_run_time = df_all["Time (s)"].mean()
                if "Median Intensity" in df_all:
                    avg_med_int = df_all["Median Intensity"].mean()
                if "Average Intensity" in df_all:
                    avg_avg_int = df_all["Average Intensity"].mean()

            # one-row dataframe for the new sheet
            agg_row = {
                "Tracy Version": __version__,
                "Pixel size (nm/px)": px_nm if px_nm is not None else "",
                "Frame time (ms)": ft_ms if ft_ms is not None else "",
                "Total movie frames": n_frames,
                "Total time (s)": total_time_s if total_time_s != "" else "",
                "Movie dimensions (px)": movie_dims_px_txt,
                "Movie dimensions (μm)": movie_dims_um_txt,
                "Total kymographs": total_kymographs,
                "Summed kymograph distances (μm)": total_kymo_distance_um,
                "Empty kymographs": empty_kymographs,
                "Number of trajectories": n_trajs,
                "Number of events (/min)": events_per_min,
                "Number of events (/um/min)": events_per_um_per_min,
                "Average net speed (μm/s)": avg_net_speed,
                "Average average speed (μm/s)": avg_avg_speed,
                "Average run length (μm)": avg_run_length,
                "Average run time (s)": avg_run_time,
                "Average median intensity": avg_med_int,
                "Average average intensity": avg_avg_int,
            }
            df_aggregate = pd.DataFrame([agg_row])

            # Write to Excel
            with pd.ExcelWriter(filename) as writer:
                df_aggregate.to_excel(writer, sheet_name="Aggregate Analysis", index=False)
                _autosize_worksheet(writer, "Aggregate Analysis", df_aggregate)
                df_data.to_excel(writer, sheet_name="Data Points", index=False)
                _autosize_worksheet(writer, "Data Points", df_data)
                df_summary.to_excel(writer, sheet_name="Per-trajectory", index=False)
                _autosize_worksheet(writer, "Per-trajectory", df_summary)
                df_per_node.to_excel(writer, sheet_name="Per-segment", index=False)
                _autosize_worksheet(writer, "Per-segment", df_per_node)
                df_per_roi.to_excel(writer, sheet_name="Per-kymograph", index=False)
                _autosize_worksheet(writer, "Per-kymograph", df_per_roi)
                for ch, df_ch in per_roi_by_channel.items():
                    df_ch.to_excel(writer, sheet_name=f"Per-kymograph Ch. {ch}", index=False)
                    _autosize_worksheet(writer, f"Per-kymograph Ch. {ch}", df_ch)
        except Exception as e:
            QMessageBox.critical(self, "Save Error", f"Failed to save trajectories: {e}")

    def hide_empty_columns(self):
        # columns to never hide, regardless of content
        always_visible = {
            "Frame A",
            "Frame B",
            "Start X,Y",
            "End X,Y",
            "Total",
            "Valid %",
            "Med. Intensity"
        }

        d_col = getattr(self.navigator, "_DIFF_D_COL", None)
        a_col = getattr(self.navigator, "_DIFF_A_COL", None)
        diff_cols = {c for c in (d_col, a_col) if c}

        for col in range(self.table_widget.columnCount()):
            hdr = self._headers[col]

            # always keep ID, any in always_visible, or any binary/value column
            ctype = self._column_types.get(hdr)
            if col == 0 or hdr in always_visible:
                self.table_widget.setColumnHidden(col, False)
                continue
            if hdr in diff_cols:
                if getattr(self.navigator, "show_diffusion", False):
                    self.table_widget.setColumnHidden(col, False)
                    continue
            elif ctype in ("binary", "value"):
                self.table_widget.setColumnHidden(col, False)
                continue

            # otherwise, only show if at least one non-empty cell
            has_value = False
            for row in range(self.table_widget.rowCount()):
                item = self.table_widget.item(row, col)
                if item and item.text().strip():
                    has_value = True
                    break
            self.table_widget.setColumnHidden(col, not has_value)

    def on_trajectory_selected_by_table(self):
        selected_rows = self.table_widget.selectionModel().selectedRows()
        if not selected_rows:
            if self.navigator is not None:
                self.navigator._ensure_traj_overlay_mode_valid(redraw=True)
            return
        index = selected_rows[0].row()
        self.on_trajectory_selected_by_index(index, zoom=True)

    def on_trajectory_selected_by_index(self, index, zoom=False):
        if index < 0 or index >= len(self.trajectories):
            return

        self.current_index = index

        # ——— locals & block redraws ———
        nav = self.navigator
        mc  = self.movieCanvas
        kc  = self.kymoCanvas
        ic  = nav.intensityCanvas
        vc  = nav.velocityCanvas

        traj = self.trajectories[index]
        ch = traj.get("channel", None)
        if ch is not None:
            nav._select_channel(ch)
        # block Qt/Matplotlib repaints
        # mc.setUpdatesEnabled(False)
        # kc.setUpdatesEnabled(False)

        try:
            # ——— 1) extract trajectory once ———
            nav.analysis_channel         = traj["channel"]
            nav.analysis_start           = traj["start"]
            nav.analysis_end             = traj["end"]
            nav.analysis_roi             = traj["roi"]
            nav.analysis_frames          = traj["frames"]
            nav.analysis_search_centers  = traj["search_centers"]
            nav.analysis_original_coords = traj["original_coords"]
            nav.analysis_intensities     = traj["intensities"]
            nav.analysis_background      = traj["background"]
            nav.analysis_fit_params      = list(zip(traj["spot_centers"],
                                                    traj["sigmas"],
                                                    traj["peaks"]))
            nav.analysis_velocities      = traj["velocities"]
            nav.analysis_trajectory_background = traj["fixed_background"]
            nav.analysis_background = traj["background"]

            # ——— 2) update the small canvases ———
            # assume these methods are fast or already optimized internally
            ic.plot_intensity(traj["frames"],
                            traj["intensities"],
                            avg_intensity=traj["average"],
                            median_intensity=traj["median"],
                            colors=nav._get_traj_colors(traj)[0])
            vc.plot_velocity_histogram(traj["velocities"])

            # ——— 3) reset looping & zoom state ———
            if self.navigator.looping:
                self.navigator.stoploop()
            mc.manual_zoom   = False
            nav.loop_index   = 0

            ic.current_index = 0
            ic.highlight_current_point()

            # ——— 4) update the analysis slider ———
            if hasattr(nav, 'analysisSlider'):
                sld = nav.analysisSlider
                sld.blockSignals(True)
                sld.setRange(0, len(nav.analysis_frames) - 1)
                sld.setValue(0)
                sld.blockSignals(False)

            # ——— 5) (Re)draw trajectories, but reuse artists ———
            if nav.traj_overlay_button.isChecked():
                mc.clear_movie_trajectory_markers()
                mc.draw_trajectories_on_movie()
                kc.clear_kymo_trajectory_markers()
                kc.draw_trajectories_on_kymo()

        finally:
            # ——— 6) single redraw and re-enable updates ———
            # mc.setUpdatesEnabled(True)
            # kc.setUpdatesEnabled(True)
            nav.jump_to_analysis_point(0, animate="ramp", zoom=zoom)
            # mc.draw_idle()
            # kc.draw_idle()

    def add_trajectory_from_navigator(self, trajid=None):
        """
        Gathers the current analysis data from the navigator and adds a new trajectory.
        """
        if (self.navigator is None or
            self.navigator.analysis_start is None or 
            self.navigator.analysis_end is None or 
            not self.navigator.analysis_frames or 
            not self.navigator.analysis_original_coords):
            # print("❌ No trajectory data found!")
            return

        channel = self.navigator.analysis_channel
        traj_background = self.navigator.analysis_trajectory_background
        start = self.navigator.analysis_start
        end = self.navigator.analysis_end
        frames = self.navigator.analysis_frames
        anchors = self.navigator.analysis_anchors
        roi = self.navigator.analysis_roi
        intensities = self.navigator.analysis_intensities
        original_coords = self.navigator.analysis_original_coords
        if self.navigator.analysis_search_centers:
            search_centers = self.navigator.analysis_search_centers
        else:
            search_centers = self.navigator.analysis_original_coords
        avg_intensity = self.navigator.analysis_avg
        median_intensity = self.navigator.analysis_median
        average_velocity = self.navigator.analysis_average_velocity
        velocities = self.navigator.analysis_velocities
        step_indices = self.navigator.analysis_step_indices
        step_medians = self.navigator.analysis_step_medians

        nodes = []
        click_source = ""
        if self.navigator.analysis_anchors and self.navigator.analysis_roi is not None:
            click_source = "kymograph"
            roi = self.navigator.analysis_roi
            for frame_idx, xk, _yk in self.navigator.analysis_anchors:
                mx, my = self.navigator.compute_roi_point(roi, xk)
                nodes.append((int(frame_idx), float(mx), float(my)))
        elif getattr(self.navigator, "analysis_points", None):
            click_source = "movie"
            for frame_idx, x, y in self.navigator.analysis_points:
                if frame_idx is None or x is None or y is None:
                    continue
                nodes.append((int(frame_idx), float(x), float(y)))

        # Store the list of refined spot centers along the trajectory.
        spot_centers = []
        sigmas = []
        peaks = []
        if hasattr(self.navigator, "analysis_fit_params"):
            for fit in self.navigator.analysis_fit_params:
                spot_centers.append(fit[0])  # This may be None if the fit failed
                sigmas.append(fit[1])
                peaks.append(fit[2])

        background = self.navigator.analysis_background if hasattr(self.navigator, "analysis_background") else None

        if trajid is None:
            trajid = self._trajectory_counter
            
        traj_data = {
            "trajectory_number": trajid,
            "channel": channel,
            "start": start,
            "end": end,
            "anchors": anchors,
            "roi": roi,
            "spot_centers": spot_centers,
            "sigmas": sigmas,
            "peaks": peaks,
            "fixed_background": traj_background,
            "background": background,
            "frames": frames,
            "original_coords": original_coords,
            "search_centers": search_centers,
            "intensities": intensities,
            "average": avg_intensity,
            "median": median_intensity,
            "velocities": velocities,
            "average_velocity": average_velocity,
            "step_indices": step_indices,
            "step_medians": step_medians,
            "nodes": nodes,
            "click_source": click_source
        } 

        traj_data["colocalization_any"]    = list(self.navigator.analysis_colocalized)
        traj_data["colocalization_by_ch"]  = {
            ch: list(flags)
            for ch, flags in self.navigator.analysis_colocalized_by_ch.items()
        }

        traj_data["motion_state"] = (
            list(getattr(self.navigator, "analysis_motion_state", []))
            if getattr(self.navigator, "analysis_motion_state", None) is not None
            else None
        )
        traj_data["motion_segments"] = (
            list(getattr(self.navigator, "analysis_motion_segments", []))
            if getattr(self.navigator, "analysis_motion_segments", None) is not None
            else None
        )

        traj_data["custom_fields"] = {}
    
        # diffusion: store if toggled
        D_COL = self.navigator._DIFF_D_COL
        A_COL = self.navigator._DIFF_A_COL
        if getattr(self.navigator, "show_diffusion", False):
            D = getattr(self.navigator, "analysis_diffusion_D", None)
            alpha = getattr(self.navigator, "analysis_diffusion_alpha", None)
            traj_data["custom_fields"][D_COL] = "" if D is None else f"{D:.4f}"
            traj_data["custom_fields"][A_COL] = "" if alpha is None else f"{alpha:.3f}"
        else:
            traj_data["custom_fields"][D_COL] = ""
            traj_data["custom_fields"][A_COL] = ""

        traj_data["segment_diffusion"] = []
        if getattr(self.navigator, "show_diffusion", False):
            try:
                traj_data["segment_diffusion"] = self._compute_segment_diffusion(
                    traj_data, self.navigator
                )
            except Exception:
                traj_data["segment_diffusion"] = []

        new = [
            (f, cx, cy)
            for f, c in zip(frames, spot_centers)
            if isinstance(c, (tuple,list)) and c[0] is not None and c[1] is not None
            for cx, cy in [c]
        ]
        self.navigator.past_centers.extend(new)

        if len(frames) == 0 or len(intensities) == 0:
            print("Error: no frames or intensities to add!")
            return

        self.trajectories.append(traj_data)

        # Assumes average_velocity is in pixels/frame (traj["average_velocity"]).
        if self.navigator.pixel_size is not None and self.navigator.frame_interval is not None and average_velocity is not None:
            velocity_nm_per_ms = (average_velocity * self.navigator.pixel_size) / self.navigator.frame_interval
            avg_vel_um_s = f"{velocity_nm_per_ms:.2f}"
            # avg_vel_um_min = f"{velocity_nm_per_ms * 60.0:.2f}"
        else:
            avg_vel_um_s = ""
            # avg_vel_um_min = ""

        num_points = len(frames)
        valid_points = sum(1 for val in intensities if val is not None and val > 0)
        percent_valid = int(100 * valid_points / num_points) if num_points > 0 else 0

        dx = end[1] - start[1]
        dy = end[2] - start[2]
        distance_px = np.hypot(dx, dy)
        time_fr = end[0]-start[0]
        distance_um_txt = ""
        time_s_txt = ""
        overall_vel_um_s_txt = ""
        if self.navigator.pixel_size is not None and self.navigator.frame_interval is not None and time_fr > 0:
            distance_um = distance_px * self.navigator.pixel_size / 1000
            time_s = time_fr * self.navigator.frame_interval / 1000
            overall_vel_um_s = distance_um/time_s
            distance_um_txt = f"{distance_um:.2f}"
            time_s_txt = f"{time_s:.2f}"
            overall_vel_um_s_txt = f"{overall_vel_um_s:.2f}"

        # Add a new row to the table.
        row = self.table_widget.rowCount()
        self.table_widget.insertRow(row)
        self.writeToTable(row, "num", str(trajid))
        if self.navigator.movieChannelCombo.count() > 1:
            self.writeToTable(row, "channel", str(channel))
        self.writeToTable(row, "startframe", str(int(start[0]) + 1))
        self.writeToTable(row, "endframe", str(int(end[0]) + 1))
        self.writeToTable(row, "startcoord", f"{start[1]:.1f}, {start[2]:.1f}")
        self.writeToTable(row, "endcoord", f"{end[1]:.1f}, {end[2]:.1f}")
        self.writeToTable(row, "distance", distance_um_txt)
        self.writeToTable(row, "time", time_s_txt)
        self.writeToTable(row, "netspeed", overall_vel_um_s_txt)  
        self.writeToTable(row, "total", str(num_points))
        self.writeToTable(row, "valid", str(percent_valid))        
        median_text = "" if median_intensity is None else f"{median_intensity:.2f}"
        avg_text = "" if avg_intensity is None else f"{avg_intensity:.2f}"
        self.writeToTable(row, "medintensity", median_text)

        if D_COL in self.custom_columns:
            self.writeToTable(row, D_COL, traj_data["custom_fields"].get(D_COL, ""))
        if A_COL in self.custom_columns:
            self.writeToTable(row, A_COL, traj_data["custom_fields"].get(A_COL, ""))

        # 3) Fill in the Channel X co. % custom columns
        # Compute one-per-trajectory % from self.analysis_colocalized.

        if self.navigator.movie.ndim == 4 and self.navigator._channel_axis is not None:
            n_chan = self.navigator.movie.shape[self.navigator._channel_axis]
            r = self.table_widget.rowCount() - 1
            ch_ref = channel
            # 1) overall percent (two-channel case)
            valid_any = [s for s in traj_data['colocalization_any'] if s is not None]
            if valid_any:
                pct_any = f"{100 * sum(1 for s in valid_any if s == 'Yes') / len(valid_any):.1f}"
            else:
                pct_any = ""

            # 2) per-channel percents (multi-channel case)
            pct_by_ch = {}
            for tgt_ch, flags in traj_data['colocalization_by_ch'].items():
                valid = [s for s in flags if s is not None]
                if valid:
                    pct_by_ch[tgt_ch] = (
                        f"{100 * sum(1 for s in valid if s == 'Yes') / len(valid):.1f}"
                    )
                else:
                    pct_by_ch[tgt_ch] = ""
            # 3) now populate
            for ch in range(1, n_chan+1):
                col_name = f"Ch. {ch} co. %"
                if ch == ch_ref:
                    val = ""
                elif n_chan == 2:
                    val = pct_any
                else:
                    val = pct_by_ch.get(ch, "")
                self._mark_custom(r, col_name, val)


        self._trajectory_counter += 1
        self.navigator.update_table_visibility()

        # Set the table's selection to the newly added trajectory.
        new_row = self.table_widget.rowCount() - 1
        self.table_widget.blockSignals(True)
        self.table_widget.selectRow(new_row)
        self.table_widget.blockSignals(False)

        # Also update the current index in the plot canvas and trigger a selection update.
        self.intensity_canvas.current_index = new_row
        self.on_trajectory_selected_by_index(new_row, zoom=True)
        if new_row == 0:
            QTimer.singleShot(0, lambda: self.navigator.jump_to_analysis_point(0, animate="discrete", zoom=True))

    def _parse_trackmate_csv(self, filename):
        df_temp = pd.read_csv(filename, header=0, engine="python")

        if "FRAME" not in df_temp.columns:
            QMessageBox.critical(self, "Error", "CSV is missing the FRAME column.")
            return None, None, None

        # Helper: check if value is numeric.
        def is_numeric(x):
            try:
                float(x)
                return True
            except (ValueError, TypeError):
                return False

        # Find first row index where the FRAME column is numeric.
        data_start = None
        for idx, val in df_temp["FRAME"].items():
            if is_numeric(val):
                data_start = idx
                break
        if data_start is None:
            QMessageBox.critical(self, "Error", "No numeric data found in the 'FRAME' column.")
            return None, None, None
        # Keep only the rows from that index onward.
        df_temp = df_temp.loc[data_start:].reset_index(drop=True)

        # Verify expected columns.
        required_csv_cols = {"TRACK_ID", "FRAME", "POSITION_X", "POSITION_Y"}
        if not required_csv_cols.issubset(set(df_temp.columns)):
            QMessageBox.critical(
                self, "Error",
                "CSV is missing one or more required columns: TRACK_ID, FRAME, POSITION_X, POSITION_Y"
            )
            return None, None, None

        # Check if the pixel size hasn't been set yet.
        if self.navigator.pixel_size is None:
            self.navigator.set_scale()  # This will open the Set Scale dialog
            # If pixel_size is still not set, bail out.
            if self.navigator.pixel_size is None:
                return None, None, None

        # Pixel size must be set before conversion.
        pixelsize = self.navigator.pixel_size
        conversion_factor = float(pixelsize) / 1000.0

        # Convert POSITION columns to numeric.
        df_temp["POSITION_X"] = pd.to_numeric(df_temp["POSITION_X"], errors="coerce")
        df_temp["POSITION_Y"] = pd.to_numeric(df_temp["POSITION_Y"], errors="coerce")
        # (If necessary, drop an extra row of NaNs)
        df_temp = df_temp.iloc[1:].reset_index(drop=True)

        # Build new DataFrame with expected column names.
        new_data = {
            "Trajectory": df_temp["TRACK_ID"],
            "Frame": pd.to_numeric(df_temp["FRAME"], errors="coerce").astype(int),
            "Original Coordinate X": df_temp["POSITION_X"] / conversion_factor,
            "Original Coordinate Y": df_temp["POSITION_Y"] / conversion_factor,
        }
        df = pd.DataFrame(new_data)
        df["Trajectory"] = pd.to_numeric(df["Trajectory"], errors="coerce")
        df["Frame"] = pd.to_numeric(df["Frame"], errors="coerce")
        df = df.sort_values(by=["Trajectory", "Frame"]).reset_index(drop=True)

        nodes_map = {}
        clicks_map = {}
        for traj_id, group in df.groupby("Trajectory"):
            group = group.sort_values("Frame")
            nodes = []
            for frame, x, y in zip(
                group["Frame"].tolist(),
                group["Original Coordinate X"].tolist(),
                group["Original Coordinate Y"].tolist()
            ):
                try:
                    frame_idx = int(frame) - 1
                    x_val = float(x)
                    y_val = float(y)
                except Exception:
                    continue
                nodes.append((frame_idx, x_val, y_val))
            nodes_map[traj_id] = nodes
            clicks_map[traj_id] = "trackmate"

        return df, nodes_map, clicks_map

    def load_trajectories(self):
        if self.navigator.movie is None:
            QMessageBox.warning(self, "", 
                "Please load a movie before loading trajectories.")
            return
        
        self.navigator.cancel_left_click_sequence()

        # Get the current movie's base name.
        movie_base = self.navigator.movieNameLabel.text()  # assuming this holds the filename, e.g. "my_movie.tif"
        # Remove the current extension and add ".csv"
        default_filename = os.path.splitext(movie_base)[0] + ".csv"
        default_filename = os.path.join(self.navigator._last_dir, default_filename)
        filename, _ = QFileDialog.getOpenFileName(
            self, 
            "Load Trajectories File", 
            default_filename,
            "Excel and CSV Files (*.xlsx *.csv)"
        )
        if not filename:
            return
        
        # try:
        ext = os.path.splitext(filename)[1].lower()

        anchors_map = {}
        roi_map     = {}
        nodes_map   = {}
        clicks_map  = {}
        segment_diff_map = {}

        self.clear_trajectories(prompt=False)
        self._custom_load_map = {}

        self.table_widget.setRowCount(0)
        self._trajectory_counter = 0

        self.show_steps = False
        self.navigator.showStepsAction.setChecked(False)

        # ----- Excel branch -----
        if ext == ".xlsx":
            xls = pd.ExcelFile(filename)
            if "Data Points" not in xls.sheet_names:
                QMessageBox.critical(self, "Error", "Sheet 'Data Points' not found.")
                return

            # --- Optionally load scale calibration from the workbook ---
            # If present, "Aggregate Analysis" contains pixel size (nm/px) and frame time (ms).
            agg_sheet = None
            for name in xls.sheet_names:
                if name.lower() == "aggregate analysis":
                    agg_sheet = name
                    break

            def _num_or_none(v):
                if v is None:
                    return None
                if isinstance(v, str) and not v.strip():
                    return None
                try:
                    if pd.isna(v):
                        return None
                except Exception:
                    pass
                try:
                    return float(v)
                except Exception:
                    return None

            loaded_px = None
            loaded_ft = None
            if agg_sheet:
                try:
                    agg_df = pd.read_excel(xls, sheet_name=agg_sheet, nrows=1)
                    if not agg_df.empty:
                        row0 = agg_df.iloc[0]
                        loaded_px = _num_or_none(row0.get("Pixel size (nm/px)", None))
                        loaded_ft = _num_or_none(row0.get("Frame time (ms)", None))
                except Exception:
                    # Ignore parse errors and continue loading.
                    pass

            cur_px = self.navigator.pixel_size
            cur_ft = self.navigator.frame_interval

            missing_px = (loaded_px is None)
            missing_ft = (loaded_ft is None)
            has_current_scale = (cur_px is not None or cur_ft is not None)

            skip_scale_apply = False
            if missing_px or missing_ft:
                if has_current_scale:
                    msg = (
                        "No pixel size and/or frame time was found in this file.\n\n"
                        f"Current pixel size: {cur_px if cur_px is not None else '—'} nm/px\n"
                        f"Current frame time: {cur_ft if cur_ft is not None else '—'} ms\n\n"
                        "Press OK to use the current scale, or Cancel to stop loading."
                    )
                    reply = QMessageBox.warning(
                        self,
                        "Missing scale",
                        msg,
                        QMessageBox.Ok | QMessageBox.Cancel,
                        QMessageBox.Ok
                    )
                    if reply == QMessageBox.Cancel:
                        return
                    skip_scale_apply = True
                else:
                    if missing_px:
                        msg = (
                            "No pixel size was found in this file.\n\n"
                            "Without a pixel size, distance and speed values may be incorrect.\n\n"
                            "Press OK to continue without applying a scale, or Cancel to stop loading."
                        )
                        reply = QMessageBox.warning(
                            self,
                            "Missing pixel size",
                            msg,
                            QMessageBox.Ok | QMessageBox.Cancel,
                            QMessageBox.Ok
                        )
                        if reply == QMessageBox.Cancel:
                            return
                        skip_scale_apply = True

            if not skip_scale_apply:
                def _different(a, b):
                    if a is None and b is None:
                        return False
                    if a is None or b is None:
                        return True
                    try:
                        return not math.isclose(float(a), float(b), rel_tol=1e-6, abs_tol=1e-6)
                    except Exception:
                        return True

                if _different(loaded_px, cur_px) or _different(loaded_ft, cur_ft):
                    msg = (
                        "The file scale differs from the current scale:\n\n"
                        f"File pixel size: {loaded_px} nm/px\n"
                        f"File frame time: {loaded_ft if loaded_ft is not None else '—'} ms\n\n"
                        f"Current pixel size: {cur_px if cur_px is not None else '—'} nm/px\n"
                        f"Current frame time: {cur_ft if cur_ft is not None else '—'} ms\n\n"
                        "Press OK to apply the file scale, or Cancel to stop loading."
                    )
                    reply = QMessageBox.warning(
                        self,
                        "Scale mismatch",
                        msg,
                        QMessageBox.Ok | QMessageBox.Cancel,
                        QMessageBox.Ok
                    )
                    if reply == QMessageBox.Cancel:
                        return

                self.navigator.pixel_size = loaded_px
                self.navigator.frame_interval = loaded_ft
                self.navigator.update_scale_label()

            if "Per-trajectory" in xls.sheet_names:
                summary_df = pd.read_excel(xls, sheet_name="Per-trajectory")
                rename_map = {
                    "Start_Frame":    "Start Frame",
                    "End_Frame":      "End Frame",
                    "Num_Points":     "Total Points",
                    "Valid_Points":   "Valid Points",
                }
                summary_df.rename(columns=rename_map, inplace=True)
                if "Kymo-Anchors" not in summary_df.columns and "Anchors" in summary_df.columns:
                    summary_df.rename(columns={"Anchors": "Kymo-Anchors"}, inplace=True)
                elif "Anchors" in summary_df.columns:
                    summary_df.drop(columns=["Anchors"], inplace=True)
                if "Movie-Anchors" not in summary_df.columns and "Nodes" in summary_df.columns:
                    summary_df.rename(columns={"Nodes": "Movie-Anchors"}, inplace=True)
                elif "Nodes" in summary_df.columns:
                    summary_df.drop(columns=["Nodes"], inplace=True)

                for _, row in summary_df.iterrows():
                    traj_id = int(row["Trajectory"])

                    # parse kymo anchors
                    a = row.get("Kymo-Anchors", "")
                    if isinstance(a, str) and a.strip():
                        try:
                            anchors_py = json.loads(a)
                        except json.JSONDecodeError:
                            anchors_py = []
                    else:
                        anchors_py = []
                    anchors_map[traj_id] = anchors_py

                    # parse ROI
                    r = row.get("ROI", "")
                    if isinstance(r, str) and r.strip():
                        try:
                            roi_clean = json.loads(r)
                            # make sure points are tuples
                            roi_clean["points"] = [
                                tuple(pt) for pt in roi_clean.get("points", [])
                            ]
                        except json.JSONDecodeError:
                            roi_clean = None
                    else:
                        roi_clean = None
                    roi_map[traj_id] = roi_clean

                    # parse Movie-Anchors (x, y, frame)
                    n = row.get("Movie-Anchors", "")
                    nodes_clean = []
                    if isinstance(n, str) and n.strip():
                        try:
                            nodes_raw = json.loads(n)
                        except json.JSONDecodeError:
                            nodes_raw = []
                        frames_raw = []
                        for item in nodes_raw:
                            if not isinstance(item, (list, tuple)) or len(item) != 3:
                                continue
                            x, y, f = item
                            try:
                                f_int = int(round(float(f)))
                            except Exception:
                                continue
                            frames_raw.append(f_int)
                            try:
                                x_f = float(x)
                                y_f = float(y)
                            except Exception:
                                continue
                            nodes_clean.append((f_int, x_f, y_f))

                        has_zero = any(f == 0 for f in frames_raw)
                        if nodes_clean:
                            nodes_clean = [
                                (int(f - 1) if (f >= 1 and not has_zero) else int(f), x, y)
                                for f, x, y in nodes_clean
                            ]
                    nodes_map[traj_id] = nodes_clean

                    clicks = row.get("Clicks", "")
                    clicks_map[traj_id] = str(clicks).strip() if pd.notna(clicks) else ""

                # Backfill Movie-Anchors from kymo anchors when missing.
                for traj_id, anchors_py in anchors_map.items():
                    if nodes_map.get(traj_id):
                        continue
                    roi = roi_map.get(traj_id)
                    if not anchors_py or roi is None:
                        continue
                    backfill = []
                    for frame_idx, xk, _yk in anchors_py:
                        try:
                            mx, my = compute_roi_point(roi, xk)
                        except Exception:
                            continue
                        backfill.append((int(frame_idx), float(mx), float(my)))
                    if backfill:
                        nodes_map[traj_id] = backfill

                # — detect any extra columns —
                known = {"Movie","Trajectory","Channel","Start Frame","End Frame",
                        "Kymo-Anchors","ROI","Clicks","Movie-Anchors","Segments",
                        "Total Points","Valid Points","Percent Valid",
                        "Search Center X Start","Search Center Y Start",
                        "Search Center X End","Search Center Y End",
                        "Distance (μm)","Time (s)","Background",
                        "Average Intensity","Median Intensity",
                        "Net Speed (px/frame)","Net Speed (μm/s)","Net Speed (μm/min)",
                        "Avg. Speed (px/frame)","Avg. Speed (μm/s)","Avg. Speed (μm/min)",
                        "Number of Steps", "Average Step Size", "Average Step Size w/Step to Background"}
                # full headers as they appear in the sheet, e.g. "Foo [value]"
                full_extra = [c for c in summary_df.columns if c not in known]

                coloc_rx = re.compile(r"^(?:Colocalized w/any channel|Ch\.\s*\d+\s*co\.\s*%)$")
                full_extra = [
                    c for c in summary_df.columns
                    if c not in known and not coloc_rx.match(c)
                ]

                def _base_name(col_name):
                    m = re.match(r"(.+)\s\[(?:binary|value)\]$", col_name)
                    return m.group(1) if m else col_name

                # drop any extras that are actually known columns with a suffix
                full_extra = [c for c in full_extra if _base_name(c) not in known]

                # build the load-map using the full names
                custom_map = {}
                for _, row in summary_df.iterrows():
                    tid = int(row["Trajectory"])
                    d = {}
                    for full in full_extra:
                        # strip off the suffix so our key is the plain name
                        m = re.match(r"(.+)\s\[(?:binary|value)\]$", full)
                        name = m.group(1) if m else full

                        val = row[full]
                        if pd.isna(val):
                            d[name] = ""
                        elif isinstance(val, float):
                            # turn 2.0 → "2", but leave 2.5 as "2.5"
                            if val.is_integer():
                                d[name] = str(int(val))
                            else:
                                d[name] = str(val)
                        else:
                            # covers ints, strings, etc.
                            d[name] = str(val)

                    custom_map[tid] = d

                # 1) figure out how many channels this movie really has:
                if self.navigator.movie is not None and self.navigator._channel_axis is not None:
                    n_chan = self.navigator.movie.shape[self.navigator._channel_axis]
                else:
                    n_chan = 1

                # 2) now prune out any stray colocalization columns that don't apply
                for traj_id, fields in custom_map.items():
                    # overall‐coloc only makes sense if exactly 2 channels
                    if "colocalization" in fields and n_chan != 2:
                        del fields["colocalization"]
                    # per‐channel coloc columns have names "Ch. {ch} co. %"
                    for name in list(fields):
                        m = re.match(r"Ch\.\s*(\d+)\s*co\.\s*%", name)
                        if m:
                            ch = int(m.group(1))
                            # drop if it refers to a channel outside 1..n_chan
                            if ch < 1 or ch > n_chan:
                                del fields[name]

                # 3) store the pruned map for later
                self._custom_load_map = custom_map

                base = [
                    "", "Channel", "Frame A", "Frame B",
                    "Start X,Y", "End X,Y",
                    "Distance μm", "Time s", "Net Speed μm/s",
                    "Total", "Valid %",
                    "Med. Intensity",
                ]

                # 1) reset in-memory header list & index map
                self._headers    = base.copy()
                self._col_index  = {hdr: i for i, hdr in enumerate(self._headers)}
                self.custom_columns = []
                self._column_types.clear()

                # 2) tell Qt about it
                self.table_widget.setColumnCount(len(self._headers))
                self.table_widget.setHorizontalHeaderLabels(self._headers)

                # parse off “[binary]” or “[value]”
                D_COL = getattr(self.navigator, "_DIFF_D_COL", None)
                A_COL = getattr(self.navigator, "_DIFF_A_COL", None)
                diff_cols = {c for c in (D_COL, A_COL) if c}

                parsed = []
                for full in full_extra:
                    m = re.match(r"(.+)\s\[(binary|value)\]$", full)
                    if m:
                        name, typ = m.group(1), m.group(2)
                    else:
                        # no suffix → assume value unless explicitly tagged
                        name, typ = full, "value"
                    parsed.append(name)
                    self._column_types[name] = typ

                # # keep any existing coloc columns
                # old_colocs = [
                #     c for c in self.custom_columns
                #     if self._column_types.get(c) == "coloc"
                # ]

                # first, clear out and re-add all user-saved fields
                self.custom_columns = []
                # (we keep _column_types around so we know each field's type)
                for name in parsed:
                    self._add_custom_column(name, col_type=self._column_types[name])

                # print("load_trajectories", "custom_columns after adding parsed", self.custom_columns)

                # now make sure every “Ch. N co. %” column is registered
                if self.navigator.movie is not None and self.navigator._channel_axis is not None:
                    n_chan = self.navigator.movie.shape[self.navigator._channel_axis]
                    for ch in range(1, n_chan+1):
                        col_name = f"Ch. {ch} co. %"
                        if col_name not in self.custom_columns: 
                            self._add_custom_column(col_name, col_type="coloc")
                else:
                    n_chan = 1

                # print("load_trajectories", "custom_columns after adding coloc", self.custom_columns)

                self.navigator._rebuild_color_by_actions()

            if "Per-segment" in xls.sheet_names:
                try:
                    seg_df = pd.read_excel(xls, sheet_name="Per-segment")
                except Exception:
                    seg_df = None

                if isinstance(seg_df, pd.DataFrame):
                    d_col = getattr(self.navigator, "_DIFF_D_COL", "D (μm²/s)")
                    a_col = getattr(self.navigator, "_DIFF_A_COL", "α")

                    def _pick_col(candidates):
                        for c in candidates:
                            if c in seg_df.columns:
                                return c
                        return None

                    d_col_use = _pick_col([d_col, "D", "Diffusion D", "Diffusion D (µm²/s)", "Diffusion D (μm²/s)"])
                    a_col_use = _pick_col([a_col, "alpha", "Alpha", "α"])

                    if "Trajectory" in seg_df.columns and "Segment" in seg_df.columns:
                        for _, row in seg_df.iterrows():
                            tid = _num_or_none(row.get("Trajectory"))
                            seg = _num_or_none(row.get("Segment"))
                            if tid is None or seg is None:
                                continue
                            try:
                                tid = int(round(float(tid)))
                                seg = int(round(float(seg)))
                            except Exception:
                                continue
                            entry = {"segment": seg}
                            if d_col_use:
                                entry["D"] = _num_or_none(row.get(d_col_use))
                            if a_col_use:
                                entry["alpha"] = _num_or_none(row.get(a_col_use))
                            if "Segment Start Frame" in seg_df.columns:
                                entry["start_frame"] = _num_or_none(row.get("Segment Start Frame"))
                            if "Segment End Frame" in seg_df.columns:
                                entry["end_frame"] = _num_or_none(row.get("Segment End Frame"))
                            segment_diff_map.setdefault(tid, []).append(entry)

            # finally read the Data Points sheet
            df = pd.read_excel(xls, sheet_name="Data Points")
        
        # ----- CSV branch -----
        elif ext == ".csv":
            df, nodes_map, clicks_map = self._parse_trackmate_csv(filename)
            if df is None:
                return
        else:
            raise Exception("Unsupported file type.")

        self.load_trajectories_from_df(
            df,
            anchors_map=anchors_map,
            roi_map=roi_map,
            nodes_map=nodes_map,
            clicks_map=clicks_map,
            segment_diffusion_map=segment_diff_map
        )

        # except Exception as e:
        #     QMessageBox.critical(self, "Load Error", f"Failed to load trajectories: {str(e)}")

        self.navigator.update_table_visibility()

        loaded_rois = [traj.get('roi') for traj in self.trajectories if isinstance(traj.get('roi'), dict)]
        if loaded_rois:
            if QMessageBox.question(
                self,
                "",
                "Generate kymographs from loaded trajectories?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            ) == QMessageBox.Yes:
                # Replay ROIs onto the kymo canvas
                self.navigator.generate_rois_from_trajectories()

                if ext == ".xlsx":
                    try:
                        if 'Per-kymograph' in xls.sheet_names:
                            pk_df = pd.read_excel(xls, sheet_name='Per-kymograph')
                            if 'ROI' in pk_df.columns:
                                pk_rois = set(pk_df['ROI'].dropna().astype(str))
                                # Determine which ROI dicts were loaded
                                loaded_rois = [traj.get('roi') for traj in self.trajectories if isinstance(traj.get('roi'), dict)]
                                loaded_jsons = set(json.dumps(roi) for roi in loaded_rois)
                                missing = pk_rois - loaded_jsons
                                if missing:
                                    if QMessageBox.question(
                                        self,
                                        "",
                                        "Load empty kymographs?",
                                        QMessageBox.Yes | QMessageBox.No,
                                        QMessageBox.Yes
                                    ) == QMessageBox.Yes:
                                        for roi_json in missing:
                                            try:
                                                roi_dict = json.loads(roi_json)
                                                self.movieCanvas.roiPoints = roi_dict.get('points', [])
                                                self.movieCanvas.finalize_roi()
                                            except json.JSONDecodeError:
                                                continue
                    except Exception as e:
                        QMessageBox.critical(
                            self, "Error",
                            f"Error: {e}."
                        )
                        return

    def load_trackmate_spots(self):
        if self.navigator.movie is None:
            QMessageBox.warning(self, "",
                "Please load a movie before loading TrackMate spots.")
            return

        self.navigator.cancel_left_click_sequence()

        movie_base = self.navigator.movieNameLabel.text()
        default_filename = os.path.splitext(movie_base)[0] + ".csv"
        default_filename = os.path.join(self.navigator._last_dir, default_filename)
        filename, _ = QFileDialog.getOpenFileName(
            self,
            "Load TrackMate spots",
            default_filename,
            "CSV Files (*.csv)"
        )
        if not filename:
            return

        ext = os.path.splitext(filename)[1].lower()
        if ext != ".csv":
            QMessageBox.critical(self, "Error", "TrackMate spots must be a .csv file.")
            return

        anchors_map = {}
        roi_map     = {}
        nodes_map   = {}
        clicks_map  = {}
        segment_diff_map = {}

        self.clear_trajectories(prompt=False)
        self._custom_load_map = {}

        self.table_widget.setRowCount(0)
        self._trajectory_counter = 0

        self.show_steps = False
        self.navigator.showStepsAction.setChecked(False)

        df, nodes_map, clicks_map = self._parse_trackmate_csv(filename)
        if df is None:
            return

        self.load_trajectories_from_df(
            df,
            anchors_map=anchors_map,
            roi_map=roi_map,
            nodes_map=nodes_map,
            clicks_map=clicks_map,
            segment_diffusion_map=segment_diff_map
        )

        self.navigator.update_table_visibility()

    def load_trajectories_from_df(self, df, anchors_map=None, roi_map=None, nodes_map=None, clicks_map=None, segment_diffusion_map=None, forcerecalc=False):

        # 1) Always require these two columns
        if "Trajectory" not in df.columns or "Frame" not in df.columns:
            QMessageBox.critical(
                self, "Error",
                "Missing columns: both 'Trajectory' and 'Frame' are required."
            )
            return

        have_coords = "Original Coordinate X" in df.columns and "Original Coordinate Y" in df.columns
        have_centers = "Search Center X" in df.columns and "Search Center Y" in df.columns

        if not have_coords and not have_centers:
            QMessageBox.critical(
                self, "Error",
                "You must provide 'Original Coordinate X'/'Original Coordinate Y' or 'Search Center X'/'Search Center Y' columns."
            )
            return
        
        if have_coords and not have_centers:
            df["Search Center X"] = df["Original Coordinate X"]
            df["Search Center Y"] = df["Original Coordinate Y"]
        elif have_centers and not have_coords:
            df["Original Coordinate X"] = df["Search Center X"]
            df["Original Coordinate Y"]  = df["Search Center Y"]

        if self.navigator.movie is not None and self.navigator._channel_axis is not None:
            n_chan = self.navigator.movie.shape[self.navigator._channel_axis]
        else:
            n_chan = 1

        selected_channel = 1

        if n_chan > 1:
            if "Channel" in df.columns:
                # the file tells us which channel each point came from;
                # gather all integers found in that column
                chans_in_file = sorted({
                    int(c) for c in df["Channel"].dropna().astype(int)
                    if 1 <= int(c) <= n_chan
                })
                if not chans_in_file:
                    QMessageBox.critical(
                        self, "Error",
                        "Found a 'Channel' column but no valid channel numbers (1–{}) in it.".format(n_chan)
                    )
                    return
            else:
                # no "Channel" column → ask the user
                chan_labels = [f"Channel {c}" for c in range(1, n_chan+1)]
                choice, ok = QInputDialog.getItem(
                    self,
                    "Select channel",
                    "Choose channel for analysis:",
                    chan_labels,
                    0,
                    False
                )
                if not ok:
                    # user canceled
                    return
                selected_channel = int(choice.split()[-1])

                self.navigator.flashchannel = False
                self.navigator._select_channel(selected_channel)
                self.navigator.flashchannel = True

        # ensure custom_load_map exists for later coloring, etc.
        if not hasattr(self, '_custom_load_map'):
            self._custom_load_map = {}

        # --- handle legacy per-point coloc columns ---
        expected_any = {"Colocalized w/any channel"}
        expected_by_ch = {f"Colocalized w/ch{ch}" for ch in range(1, n_chan+1)}
        expected_cols = expected_any | expected_by_ch
        present = set(df.columns) & expected_cols
        if present and present != expected_cols:
            missing = expected_cols - present
            QMessageBox.warning(
                self, "Colocalization columns",
                "Trajectory file has a partial set of colocalization columns;\n"
                f"missing {missing}. They will be dropped."
            )
            df = df.drop(columns=list(present), errors='ignore')

        # Determine whether intensity, sigma, and peak columns exist.
        intensity_exists = "Intensity" in df.columns
        sigma_exists = "Sigma" in df.columns
        peak_exists = "Peak" in df.columns
        background_exists = "Background" in df.columns

        if anchors_map is None:
            anchors_map = {}
        if roi_map is None:
            roi_map = {}
        if nodes_map is None:
            nodes_map = {}
        if clicks_map is None:
            clicks_map = {}
        if segment_diffusion_map is None:
            segment_diffusion_map = {}

        fixed_background = None

        # Count how many trajectories need recalculation
        trajectories_need_recalc = 0
        for traj_num, group in df.groupby("Trajectory"):
            group = group.sort_values("Frame")
            frames = group["Frame"].tolist()
            expected_frames = list(range(min(frames), max(frames) + 1))
            if (expected_frames != frames
                or not intensity_exists
                or not sigma_exists
                or not peak_exists
                or not background_exists):
                trajectories_need_recalc += 1

        if not forcerecalc:
            recalc_mode = False

        if trajectories_need_recalc > 0:
            # Instead of a simple question, show the custom dialog.
            current_mode = self.navigator.tracking_mode if hasattr(self.navigator, "tracking_mode") else "Independent"
            current_radius = self.navigator.searchWindowSpin.value()
            message = (f"{trajectories_need_recalc} trajectory(ies) have missing points or spot parameters.")
            recalc_dialog = RecalcDialog(current_mode, current_radius, message=message, parent=self)
            result = recalc_dialog.exec_()
            if result != QDialog.Accepted:
                return
            # Retrieve new selections from the dialog.
            new_mode = recalc_dialog.new_mode
            new_radius = recalc_dialog.new_radius
            # Update the GUI state.
            self.navigator.searchWindowSpin.setValue(new_radius)
            self.navigator.tracking_mode = new_mode
            if hasattr(self.navigator, "trackingModeCombo"):
                self.trackingModeCombo.setCurrentText(new_mode)

            recalc_mode = True

        # --- Process each trajectory ---
        # Create a list of groups so we can report progress.
        groups = list(df.groupby("Trajectory"))
        progress = QProgressDialog("Processing...", "Cancel", 0, len(groups), self)
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.setAutoClose(False)
        progress.setAutoReset(False)
        progress.show()

        loaded = []
        self.navigator.past_centers = []

        force_cancel = False

        for t, (traj_num, group) in enumerate(groups):
            if progress.wasCanceled() or force_cancel:
                break

            # —— extract & validate channel column ——
            if "Channel" in group.columns:
                chans = set(group["Channel"].tolist())
                if len(chans) == 1:
                    channel = chans.pop()
                else:
                    # warn but don’t bail out
                    sorted_chans = sorted(chans)
                    QMessageBox.warning(
                        self,
                        "Ambiguous Channel",
                        f"Trajectory {traj_num} has multiple channels {sorted_chans}.\n"
                        f"Defaulting to {sorted_chans[0]}."
                    )
                    channel = sorted_chans[0]
            else:
                channel = selected_channel

            if "Step Number" in group.columns and "Step Intensity Value" in group.columns:

                step_rows = group.copy()
                step_rows["Step Number"] = pd.to_numeric(step_rows["Step Number"], errors="coerce")
                step_rows["Step Intensity Value"] = pd.to_numeric(
                    step_rows["Step Intensity Value"],
                    errors="coerce"
                )
                step_rows = step_rows.dropna(subset=["Step Number", "Step Intensity Value"])
                segs = []
                for sn, segdf in step_rows.groupby("Step Number"):
                    sn = int(sn)
                    frs = segdf["Frame"].astype(int).tolist()
                    start_f = min(frs)
                    end_f   = max(frs)
                    median  = float(segdf["Step Intensity Value"].iloc[0])
                    segs.append((start_f, end_f, median))
                segs.sort(key=lambda x: x[0])
                step_idxs = [s[0] for s in segs]

                self.navigator.showStepsAction.blockSignals(True)
                self.navigator.showStepsAction.setChecked(True)
                self.navigator.show_steps = True
                self.navigator.showStepsAction.blockSignals(False)

            else:
                segs = None
                step_idxs = None

            # build the names we expect
            expected_any   = {"Colocalized w/any channel"}
            expected_by_ch = {f"Colocalized w/ch{ch}" for ch in range(1, n_chan+1)}
            # (we'll ignore the one for the reference‐channel itself later)
            expected_cols = expected_any | expected_by_ch

            # find which of these actually live in df
            present = set(df.columns) & expected_cols

            if present:
                # if *any* coloc columns are present, require *all* of them
                missing = expected_cols - present
                if missing:
                    QMessageBox.warning(
                        self, "Colocalization columns",
                        "Trajectory file has some colocalization columns but not the full set\n"
                        f"({missing}). They will be dropped."
                    )
                    # drop any partials
                    df = df.drop(columns=list(present), errors='ignore')

            group = group.sort_values("Frame")
            anchors = anchors_map.get(traj_num, [])
            roi     = roi_map.get(traj_num, None)
            frames = [x-1 for x in group["Frame"].tolist()]

            # fallback to the DataFrame’s Search Centers
            x_coords = group["Original Coordinate X"].tolist()
            y_coords = group["Original Coordinate Y"].tolist()
            points = [
                (frame, float(x), float(y))
                for frame, x, y in zip(frames, x_coords, y_coords)
            ]

            if recalc_mode:
                
                try:
                    # --- Recalculation branch ---
                    frames = [f - 1 for f in group["Frame"].tolist()]
                    full_frames = list(range(min(frames), max(frames) + 1))

                    intensities = []
                    spot_centers = []
                    sigmas = []
                    peaks = []
                    background = []
                    fixed_background = None

                    self.navigator._select_channel(channel)

                    # 2) check for a consistent “yes” background flag
                    if "Background from trajectory" in group:
                        flags = (
                            group["Background from trajectory"]
                            .dropna()
                            .astype(str)
                            .str.strip()
                            .str.lower()
                        )
                        if not flags.empty and flags.eq("yes").all():
                            # pull all numeric backgrounds in this group
                            if "Background" in group:
                                uniq_bg = group["Background"].dropna().unique()
                                if len(uniq_bg) == 1:
                                    fixed_background = float(uniq_bg[0])
                                else:
                                    # ambiguous numeric values → discard
                                    fixed_background = None

                    if len(anchors) >= 2 and roi is not None:
                        points = []
                        for i in range(len(anchors) - 1):
                            f1, xk1, _ = anchors[i]
                            f2, xk2, _ = anchors[i + 1]

                            # include f1 on the first segment, then skip it
                            seg = list(range(f1, f2 + 1)) if i == 0 else list(range(f1+1, f2+1))
                            n   = len(seg)
                            xs  = np.linspace(xk1, xk2, n, endpoint=True)

                            for j, f in enumerate(seg):
                                xk = xs[j]
                                mx, my = compute_roi_point(roi, xk)
                                points.append((f, mx, my))

                        # Unpack pts into downstream variables.
                        frames_used = [p[0] for p in points]
                        traj_coords = [(p[1], p[2]) for p in points]

                    if fixed_background is None:
                        fixed_background = self.navigator.compute_trajectory_background(
                            self.navigator.get_movie_frame,
                            points,
                            crop_size=int(2 * self.navigator.searchWindowSpin.value())
                        )

                        # Call the analysis function.
                    frames_used, traj_coords, traj_centers, intensities, fit_params, background = self.navigator._compute_analysis(points, bg=fixed_background, showprogress=False)

                    spot_centers = [fp[0] for fp in fit_params]
                    sigmas = [fp[1] for fp in fit_params]
                    peaks = [fp[2] for fp in fit_params]

                    # You can use frames_used and traj_coords (which are the refined trajectory coordinates)
                    # to calculate additional measures (e.g. velocities) if needed.
                    start = (full_frames[0], x_coords[0], y_coords[0])
                    end = (full_frames[-1], x_coords[-1], y_coords[-1])
                    frames_used = full_frames
                    valid_ints = [val for val, spot in zip(intensities, spot_centers)
                                if val is not None and val > 0 and spot is not None]
                    avg = float(np.mean(valid_ints)) if valid_ints else None
                    med = float(np.median(valid_ints)) if valid_ints else None
                    velocities = []
                    for i in range(1, len(full_frames)):
                        if spot_centers[i-1] is None or spot_centers[i] is None:
                            velocities.append(None)
                        else:
                            dx = spot_centers[i][0] - spot_centers[i-1][0]
                            dy = spot_centers[i][1] - spot_centers[i-1][1]
                            velocities.append(np.hypot(dx, dy))

                    if velocities:
                        valid_vels = [v for v in velocities if v is not None]
                        average_velocity = float(np.mean(valid_vels)) if valid_vels else None
                    else:
                        average_velocity = None

                except Exception as e:
                    QMessageBox.critical(self, "", f"Loading failed: {str(e)}")
                    force_cancel = True

            else:
                sigmas = group["Sigma"].tolist() if "Sigma" in group.columns else []
                peaks = group["Peak"].tolist() if "Peak" in group.columns else []
                background = group["Background"].tolist() if "Background" in group.columns else []
                if "Intensity" in group.columns:
                    # Replace NaN with None
                    intensities = [None if pd.isna(val) else val for val in group["Intensity"].tolist()]
                else:
                    intensities = []
                valid_ints = [val for val in intensities if val is not None and val > 0]
                avg = float(np.mean(valid_ints)) if valid_ints else None
                med = float(np.median(valid_ints)) if valid_ints else None
                row_first = group.iloc[0]
                row_last = group.iloc[-1]
                start = (frames[0], float(row_first["Original Coordinate X"]), float(row_first["Original Coordinate Y"]))
                end = (frames[-1], float(row_last["Original Coordinate X"]), float(row_last["Original Coordinate Y"]))
                if "Spot Center X" in group.columns and "Spot Center Y" in group.columns:
                    spot_centers = []
                    for x, y in zip(group["Spot Center X"], group["Spot Center Y"]):
                        sx = float(x) if pd.notnull(x) and x != "" else None
                        sy = float(y) if pd.notnull(y) and y != "" else None
                        spot_centers.append((sx, sy) if sx is not None and sy is not None else None)
                else:
                    spot_centers = []
                traj_coords = list(zip(group["Original Coordinate X"], group["Original Coordinate Y"]))
                try:
                    traj_centers = list(zip(group["Search Center X"], group["Search Center Y"]))
                except:
                    traj_centers = traj_coords
                frames_used = frames

                # 2) check for a consistent “yes” background flag
                if "Background from trajectory" in group:
                    flags = (
                        group["Background from trajectory"]
                        .dropna()
                        .astype(str)
                        .str.strip()
                        .str.lower()
                    )
                    if not flags.empty and flags.eq("yes").all():
                        # pull all numeric backgrounds in this group
                        if "Background" in group:
                            uniq_bg = group["Background"].dropna().unique()
                            if len(uniq_bg) == 1:
                                fixed_background = float(uniq_bg[0])
                            else:
                                # ambiguous numeric values → discard
                                fixed_background = None

                average_velocity = None
                # Attempt to retrieve velocities from the imported file.
                if "Speed (px/frame)" in group.columns:
                    velocities = group["Speed (px/frame)"].tolist()
                else:
                    # Recalculate velocities based on spot centers.
                    velocities = []
                    for i in range(1, len(frames_used)):
                        if spot_centers[i-1] is None or spot_centers[i] is None:
                            velocities.append(None)
                        else:
                            dx = spot_centers[i][0] - spot_centers[i-1][0]
                            dy = spot_centers[i][1] - spot_centers[i-1][1]
                            velocities.append(np.hypot(dx, dy))
                if velocities:
                    average_velocity = float(group["Speed (px/frame)"].mean()) 

            traj = {
                "trajectory_number": int(traj_num),
                "channel": channel,
                "start": start,
                "end": end,
                "anchors": anchors,
                "roi": roi,
                "spot_centers": spot_centers,
                "sigmas": sigmas,
                "peaks": peaks,
                "fixed_background": fixed_background,
                "background": background,
                "frames": frames_used,
                "original_coords": traj_coords,
                "search_centers": traj_centers,
                "intensities": intensities,
                "average": avg,
                "median": med,
                "velocities": velocities,
                "average_velocity": average_velocity,
                "step_indices": step_idxs,
                "step_medians":  segs,
                "nodes": nodes_map.get(traj_num, []),
                "click_source": clicks_map.get(traj_num, ""),
            }
            traj["segment_diffusion"] = segment_diffusion_map.get(int(traj_num), [])

            # --- load per-point colocalization flags if present ---
            n_pts = len(frames_used)
            # overall-any flags
            if "Colocalized w/any channel" in group.columns:
                traj["colocalization_any"] = [
                    v if v in ("Yes", "No") else None
                    for v in group["Colocalized w/any channel"].fillna("").tolist()
                ]
            else:
                traj["colocalization_any"] = [None] * n_pts

            # per-channel flags
            by_ch = {}
            for ch in range(1, n_chan+1):
                if ch == channel:
                    continue
                col = f"Colocalized w/ch{ch}"
                if col in group.columns:
                    flags = group[col].fillna("").tolist()
                    by_ch[ch] = [v if v in ("Yes", "No") else None for v in flags]
                else:
                    by_ch[ch] = [None] * n_pts
            traj["colocalization_by_ch"] = by_ch

            # 3) compute the per‐trajectory percentages
            computed_cf = {}
            valid_any = [v for v in traj["colocalization_any"] if v is not None]
            pct_any   = f"{100*sum(1 for s in valid_any if s=='Yes')/len(valid_any):.1f}" if valid_any else ""

            for ch in range(1, n_chan+1):
                col_name = f"Ch. {ch} co. %"
                if ch == channel:
                    computed_cf[col_name] = ""
                elif n_chan == 2:
                    computed_cf[col_name] = pct_any
                else:
                    flags = traj["colocalization_by_ch"].get(ch, [])
                    valid = [v for v in flags if v is not None]
                    computed_cf[col_name] = (
                        f"{100*sum(1 for s in valid if s=='Yes')/len(valid):.1f}"
                        if valid else ""
                    )

            # 4) merge in any other custom fields the user saved
            loaded_cf = self._custom_load_map.get(traj_num, {})
            merged_cf = {**loaded_cf, **computed_cf}
            traj["custom_fields"] = merged_cf

            # print("traj[custom_fields] after load from df", traj["custom_fields"])

            new = [
                (f, cx, cy)
                for f, c in zip(frames_used, spot_centers)
                if isinstance(c, (tuple,list)) and c[0] is not None and c[1] is not None
                for cx, cy in [c]
            ]
            self.navigator.past_centers.extend(new)

            loaded.append(traj)
            progress.setValue(t+1)
            QApplication.processEvents()

        progress.close()

        if force_cancel:
            return

        self.trajectories = loaded

        self.table_widget.setRowCount(0)

        for traj in loaded:

            avg_vel_um_s_txt = ""
            if self.navigator.pixel_size is not None and self.navigator.frame_interval is not None and traj["average_velocity"] is not None:
                velocity_nm_per_ms = (traj["average_velocity"] * self.navigator.pixel_size) / self.navigator.frame_interval
                avg_vel_um_s_txt = f"{velocity_nm_per_ms:.2f}"

            dx = traj['end'][1] - traj['start'][1]
            dy = traj['end'][2] - traj['start'][2]
            distance_px = np.hypot(dx, dy)
            time_fr = traj['end'][0]-traj['start'][0]
            distance_um_txt = ""
            time_s_txt = ""
            overall_vel_um_s_txt = ""
            if self.navigator.pixel_size is not None and self.navigator.frame_interval is not None and time_fr > 0:
                distance_um = distance_px * self.navigator.pixel_size / 1000
                time_s = time_fr * self.navigator.frame_interval / 1000
                overall_vel_um_s = distance_um/time_s
                distance_um_txt = f"{distance_um:.2f}"
                time_s_txt = f"{time_s:.2f}"
                overall_vel_um_s_txt = f"{overall_vel_um_s:.2f}"

            num_points = len(traj['frames'])
            valid_points = sum(1 for val in traj['intensities'] if val is not None and val > 0)
            percent_valid = int(100 * valid_points / num_points) if num_points > 0 else 0

            ch = traj.get("channel")
            channel_str = str(ch) if ch is not None else ""

            row = self.table_widget.rowCount()
            
            self.table_widget.insertRow(row)
            self.writeToTable(row, "num", str(traj["trajectory_number"]))
            if self.navigator.movieChannelCombo.count() > 1:
                self.writeToTable(row, "channel", channel_str)
            self.writeToTable(row, "startframe", str(int(traj["start"][0]) + 1))
            self.writeToTable(row, "endframe",   str(int(traj["end"][0])   + 1))
            start_xy = f"{traj['start'][1]:.1f}, {traj['start'][2]:.1f}" if traj['start'][1] is not None else "N/A"
            end_xy   = f"{traj['end'][1]:.1f}, {traj['end'][2]:.1f}" if traj['end'][1] is not None else "N/A"
            self.writeToTable(row, "startcoord", start_xy)
            self.writeToTable(row, "endcoord", end_xy)
            self.writeToTable(row, "distance", distance_um_txt)
            self.writeToTable(row, "time", time_s_txt)
            self.writeToTable(row, "netspeed", overall_vel_um_s_txt)  
            self.writeToTable(row, "total", str(num_points))
            self.writeToTable(row, "valid", str(percent_valid))
            median_text = "" if traj['median'] is None else f"{traj['median']:.2f}"
            avg_text = "" if traj['average'] is None else f"{traj['average']:.2f}"
            self.writeToTable(row, "medintensity", median_text)

            # print("custom_columns before writing to table", self.custom_columns)

            for col in self.custom_columns:
                val = traj.get("custom_fields", {}).get(col, "")
                # print("write to table at the end of load from df: col, val", col, val)
                if val is None or (isinstance(val, float) and math.isnan(val)) or pd.isna(val):
                    val = ""
                # Make sure the column really exists in the table:
                if col in self.custom_columns:
                    self.writeToTable(row, col, str(val))

        self._trajectory_counter = max(traj["trajectory_number"] for traj in loaded) + 1

        if self.table_widget.rowCount() > 0:
            self.table_widget.selectRow(0)

        self.movieCanvas.draw_trajectories_on_movie()
        self.kymoCanvas.draw_trajectories_on_kymo()

        self.movieCanvas.draw()
        self.kymoCanvas.draw()

    def updateTableRow(self, row, traj):
        """
        1) replace the dict in self.trajectories[row]
        2) clear that row in the QTableWidget
        3) recompute summary fields and write them back
        """
        try:
            self.trajectories[row] = traj
        except Exception as e:
            QMessageBox.critical(self, "", f"Error: {str(e)}")
            return
        # clear old items
        for col in range(self.table_widget.columnCount()):
            self.table_widget.takeItem(row, col)

        # grab everything back out
        start = traj["start"]
        end   = traj["end"]
        frames = traj["frames"]
        ints   = traj["intensities"]
        avg_int = traj["average"]
        med_int = traj["median"]
        avg_vel_pf = traj["average_velocity"]

        # compute μm distance / time / speed
        dx = end[1] - start[1]; dy = end[2] - start[2]
        px_dist = np.hypot(dx, dy)
        dt_fr   = end[0] - start[0]

        if (self.navigator.pixel_size is not None and
            self.navigator.frame_interval is not None and
            dt_fr > 0):
            μm_dist = px_dist * self.navigator.pixel_size / 1000
            secs    = dt_fr * self.navigator.frame_interval / 1000
            net_sp  = μm_dist / secs
            dist_txt   = f"{μm_dist:.2f}"
            time_txt   = f"{secs:.2f}"
            netspeed_txt = f"{net_sp:.2f}"
        else:
            dist_txt = time_txt = netspeed_txt = ""

        total_pts = len(frames)
        valid_pts = sum(1 for v in ints if v and v>0)
        valid_pct = int(100 * valid_pts / total_pts) if total_pts else 0

        avg_txt = "" if avg_int is None else f"{avg_int:.2f}"
        med_txt = "" if med_int is None else f"{med_int:.2f}"
        avvel_txt = ""
        if (self.navigator.pixel_size is not None and
            self.navigator.frame_interval is not None and
            avg_vel_pf is not None):
            μmps = (avg_vel_pf * self.navigator.pixel_size) / self.navigator.frame_interval
            avvel_txt = f"{μmps:.2f}"

        ch = traj.get("channel")
        channel_str = str(ch) if ch is not None else ""

        # Write back same columns as add_trajectory.
        self.writeToTable(row, "num",          str(traj["trajectory_number"]))
        if self.navigator.movieChannelCombo.count() > 1:
            self.writeToTable(row, "channel", channel_str)
        self.writeToTable(row, "startframe",   str(int(start[0]) + 1))
        self.writeToTable(row, "endframe",     str(int(end[0])   + 1))
        self.writeToTable(row, "startcoord",   f"{start[1]:.1f}, {start[2]:.1f}")
        self.writeToTable(row, "endcoord",     f"{end[1]:.1f}, {end[2]:.1f}")
        self.writeToTable(row, "distance",     dist_txt)
        self.writeToTable(row, "time",         time_txt)
        self.writeToTable(row, "netspeed",     netspeed_txt)
        self.writeToTable(row, "total",        str(total_pts))
        self.writeToTable(row, "valid",        str(valid_pct))
        self.writeToTable(row, "medintensity", med_txt)

        for col in self.custom_columns:
            # get the saved string (or "")
            val = traj.get("custom_fields", {}).get(col, "")
            # write it back into the table
            self.writeToTable(row, col, str(val))

    def shortcut_recalculate(self):
        selected = self.table_widget.selectionModel().selectedRows()
        if len(selected) > 1:
            reply = QMessageBox.question(
                self, "Recalculate Trajectories",
                "Recalculate selected trajectories?",
                QMessageBox.Ok | QMessageBox.Cancel,
                QMessageBox.Ok
            )
            if reply != QMessageBox.Ok:
                return

        # no dialog, just do it
        self.recalculate_trajectory(prompt=False)
        self.navigator.flash_message("Recalculated")

    def recalculate_trajectory(self, prompt=True):
        rows = [idx.row() for idx in self.table_widget.selectionModel().selectedRows()]
        if not rows:
            QMessageBox.warning(self, "", "Select at least one trajectory to recalculate")
            return

        # backup originals
        originals = {r: copy.deepcopy(self.trajectories[r]) for r in rows}
        # remove old centers in whoever is selected
        for r in rows:
            old = self.trajectories[r]
            centers_to_remove = [
                (f, cx, cy)
                for f, c in zip(old["frames"], old["search_centers"])
                if isinstance(c, (tuple, list)) and len(c) == 2 and c[0] is not None
                for cx, cy in [c]
            ]
            self.navigator._remove_past_centers(centers_to_remove)

        if len(rows) == 1:
            row = rows[0]
            new_traj = self._rebuild_one_trajectory(self.trajectories[row], self.navigator)
            self.trajectories[row] = new_traj
            self.updateTableRow(row, new_traj)
            if self.navigator.traj_overlay_button.isChecked():
                self.on_trajectory_selected_by_index(row)
            return

        # else: multi‐trajectory → pop up dialog if requested, then spawn worker
        if prompt:
            mode = getattr(self.navigator, "tracking_mode", "Independent")
            rad  = self.navigator.searchWindowSpin.value()
            dlg  = RecalcDialog(mode, rad,
                                message=f"{len(rows)} trajectories need recalc",
                                parent=self)
            if dlg.exec_() != QDialog.Accepted:
                return
            self.navigator.searchWindowSpin.setValue(dlg.new_radius)
            self.navigator.tracking_mode = dlg.new_mode
            if hasattr(self, "trackingModeCombo"):
                self.trackingModeCombo.setCurrentText(dlg.new_mode)

        total_frames = sum(len(self.trajectories[r]["frames"]) for r in rows)
        master = QProgressDialog("Recalculating", "Cancel", 0, total_frames, self)
        master.setWindowModality(Qt.WindowModal)
        master.setMinimumDuration(0)
        master.show()

        self.navigator._suppress_internal_progress = True

        worker = RecalcWorker(rows, self.trajectories, self.navigator)
        thread = QThread(self)
        worker.moveToThread(thread)
        self._recalc_thread = thread
        self._recalc_worker = worker

        worker.progress.connect(master.setValue)
        master.canceled.connect(worker.cancel)

        def on_finished(results):
            for row, new_traj in results:
                self.trajectories[row] = new_traj
                self.updateTableRow(row, new_traj)
            cleanup()

        worker.finished.connect(on_finished)

        def on_canceled():
            for r, orig in originals.items():
                self.trajectories[r] = orig
                self.updateTableRow(r, orig)
            cleanup()

        worker.canceled.connect(on_canceled)

        def cleanup():
            master.close()
            self.navigator._suppress_internal_progress = False
            thread.quit()
            thread.wait()
            if self.navigator.traj_overlay_button.isChecked() and rows:
                self.on_trajectory_selected_by_index(rows[0])
            self.navigator._refresh_intensity_canvas()

        thread.started.connect(worker.run)
        thread.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        thread.start()

    def _compute_segment_diffusion(self, traj: dict, navigator) -> List[dict]:
        nodes = traj.get("nodes") or []
        anchors = traj.get("anchors") or []
        roi = traj.get("roi", None)
        if not nodes and anchors and roi is not None:
            for frame, ax_x, _ax_y in anchors:
                mx, my = navigator.compute_roi_point(roi, ax_x)
                nodes.append((int(frame), float(mx), float(my)))

        nodes_sorted = [
            n for n in nodes
            if isinstance(n, (list, tuple)) and len(n) >= 3
        ]
        nodes_sorted.sort(key=lambda n: n[0])
        if len(nodes_sorted) < 2:
            return []

        if navigator.pixel_size is None or navigator.frame_interval is None:
            return []

        frames = traj.get("frames", []) or []
        spots = traj.get("spot_centers", []) or []
        out = []
        for idx in range(len(nodes_sorted) - 1):
            start_frame = int(nodes_sorted[idx][0])
            end_frame = int(nodes_sorted[idx + 1][0])
            seg_start = start_frame if idx == 0 else (start_frame + 1)
            seg_indices = [
                i for i, f in enumerate(frames)
                if seg_start <= f <= end_frame
            ]
            seg_frames = [frames[i] for i in seg_indices]
            seg_spots = [
                spots[i] if i < len(spots) else None
                for i in seg_indices
            ]
            try:
                D, alpha = navigator.compute_diffusion_for_data(seg_frames, seg_spots)
            except ValueError:
                D, alpha = (None, None)
            out.append({
                "segment": idx + 1,
                "start_frame": start_frame,
                "end_frame": end_frame,
                "D": D,
                "alpha": alpha,
            })
        return out

    @staticmethod
    def _rebuild_one_trajectory(old: dict, navigator) -> dict:
        """
        Given an existing trajectory‐dict `old`, recompute everything (fits, intensities,
        velocities, colocalization, steps, custom_fields) and return a brand‐new traj_data dict.
        """
        # 1) build pts
        anchors, roi = old["anchors"], old["roi"]
        nodes = old.get("nodes", [])
        if len(anchors) > 1 and roi is not None:
            pts = []
            for i in range(len(anchors) - 1):
                f1, x1, _ = anchors[i]
                f2, x2, _ = anchors[i+1]
                seg = range(f1, f2+1) if i == 0 else range(f1+1, f2+1)
                xs  = np.linspace(x1, x2, len(seg), endpoint=True)
                for j, f in enumerate(seg):
                    mx, my = navigator.compute_roi_point(roi, xs[j])
                    pts.append((f, mx, my))
        elif len(nodes) > 1:
            pts = [(f, x, y) for f, x, y in nodes]
        else:
            pts = [(f, x, y) for f, (x, y) in zip(old["frames"], old["original_coords"])]

        # 2) recompute background + analysis
        trajectory_background = navigator.compute_trajectory_background(
            navigator.get_movie_frame, pts,
            crop_size=int(2 * navigator.searchWindowSpin.value())
        )
        # showprogress=True here; batch recalc suppresses internal progress via _suppress_internal_progress.
        frames, coords, centers, ints, fit, background = navigator._compute_analysis(
            pts,
            trajectory_background,
            showprogress=True
        )

        # 3) unpack fits → spots, sigmas, peaks; compute avg/med intensities
        spots  = [p[0] for p in fit]
        sigmas = [p[1] for p in fit]
        peaks  = [p[2] for p in fit]
        valid_ints = [v for v, s in zip(ints, spots) if v and v > 0 and s]
        avg_int = float(np.mean(valid_ints)) if valid_ints else None
        med_int = float(np.median(valid_ints)) if valid_ints else None

        # 4) recalc velocities (pixel/frame)
        vels = []
        for i in range(1, len(spots)):
            p0, p1 = spots[i-1], spots[i]
            if p0 is None or p1 is None:
                vels.append(None)
            else:
                vels.append(np.hypot(p1[0] - p0[0], p1[1] - p0[1]))
        good_vels = [v for v in vels if v is not None]
        avg_vpf   = float(np.mean(good_vels)) if good_vels else None

        # 5) assemble the new trajectory dict (colocalization & steps to follow)
        traj_nodes = list(nodes)
        if len(anchors) > 1 and roi is not None:
            traj_nodes = []
            for frame, ax_x, _ax_y in anchors:
                mx, my = navigator.compute_roi_point(roi, ax_x)
                traj_nodes.append((int(frame), float(mx), float(my)))

        start = pts[0] if pts else old["start"]
        end = pts[-1] if pts else old["end"]

        traj_data = {
            "trajectory_number": old["trajectory_number"],
            "channel":           old["channel"],
            "start":             start,
            "end":               end,
            "anchors":           anchors,
            "roi":               roi,
            "spot_centers":      spots,
            "sigmas":            sigmas,
            "peaks":             peaks,
            "fixed_background":  trajectory_background,
            "background":        background,
            "frames":            frames,
            "original_coords":   coords,
            "search_centers":    centers,
            "intensities":       ints,
            "average":           avg_int,
            "median":            med_int,
            "velocities":        vels,
            "average_velocity":  avg_vpf,
            "nodes":             traj_nodes,
            "click_source":      old.get("click_source", "")
        }

        # 6) copy over existing custom_fields (so we don’t lose any manual edits)
        traj_data["custom_fields"] = old.get("custom_fields", {}).copy()
        traj_data["segment_diffusion"] = old.get("segment_diffusion", [])

        # 7) colocalization
        N = len(frames)
        if getattr(navigator, "check_colocalization", False) and navigator.movie.ndim == 4:
            # overall ANY‐channel
            navigator.analysis_frames     = frames
            navigator.analysis_fit_params = list(zip(spots, sigmas, peaks))
            navigator.analysis_channel    = old["channel"]
            navigator._compute_colocalization(showprogress=False)
            any_list = list(navigator.analysis_colocalized)

            # per‐channel flags if >2
            by_ch = {}
            n_chan = navigator.movie.shape[navigator._channel_axis]
            for tgt in range(1, n_chan+1):
                if tgt == old["channel"]:
                    continue
                navigator.analysis_channel = tgt
                navigator._compute_colocalization(showprogress=False)
                by_ch[tgt] = list(navigator.analysis_colocalized)
        else:
            any_list = [None]*N
            by_ch    = {
                ch: [None]*N
                for ch in range(1, (navigator.movie.shape[navigator._channel_axis] if navigator._channel_axis is not None else 1)+1)
                if ch != old["channel"]
            }

        traj_data["colocalization_any"]   = any_list
        traj_data["colocalization_by_ch"] = by_ch

        # 8) populate per‐trajectory % co‐loc in custom_fields
        n_chan = (navigator.movie.shape[navigator._channel_axis] if navigator._channel_axis is not None else 1)
        ref_ch = old["channel"]
        cf     = traj_data.setdefault("custom_fields", {})
        if n_chan == 2:
            valid = [s for s in any_list if s is not None]
            pct   = f"{100*sum(1 for s in valid if s=='Yes')/len(valid):.1f}" if valid else ""
            for ch in (1, 2):
                cf[f"Ch. {ch} co. %"] = "" if ch == ref_ch else pct
        else:
            for ch in range(1, n_chan+1):
                name = f"Ch. {ch} co. %"
                if ch == ref_ch:
                    cf[name] = ""
                else:
                    flags = by_ch.get(ch, [])
                    valid = [s for s in flags if s is not None]
                    cf[name] = (
                        f"{100*sum(1 for s in valid if s=='Yes')/len(valid):.1f}"
                        if valid else ""
                    )

        # 9) steps (only if show_steps=True)
        if getattr(navigator, "show_steps", False):
            idxs, meds = navigator.compute_steps_for_data(frames, ints)
            traj_data["step_indices"] = idxs
            traj_data["step_medians"] = meds
        else:
            traj_data["step_indices"] = None
            traj_data["step_medians"] = None

        if getattr(navigator, "show_diffusion", False):
            # Require physical scale for diffusion (no px/frame outputs).
            if navigator.pixel_size is None or navigator.frame_interval is None:
                raise ValueError(
                    "Diffusion requires scale: set Pixel size (nm) and Frame interval (ms) before recalculating."
                )

            D, alpha = navigator.compute_diffusion_for_data(
                traj_data["frames"],
                traj_data["spot_centers"],
            )

            cf = traj_data.setdefault("custom_fields", {})
            d_col = getattr(navigator, "_DIFF_D_COL", "Diffusion D (µm²/s)")
            a_col = getattr(navigator, "_DIFF_A_COL", "Diffusion α")
            cf[d_col] = "" if D is None else f"{D:.4f}"
            cf[a_col] = "" if alpha is None else f"{alpha:.3f}"
            try:
                traj_data["segment_diffusion"] = navigator.trajectoryCanvas._compute_segment_diffusion(
                    traj_data, navigator
                )
            except Exception:
                traj_data["segment_diffusion"] = []
        # If diffusion display is off, do NOT overwrite any existing saved values.

        return traj_data

    def _recalculate_trajectory_legacy(self, prompt=True):
        # 1) gather selection
        rows = [idx.row() for idx in self.table_widget.selectionModel().selectedRows()]
        if not rows:
            QMessageBox.warning(self, "", "Select at least one trajectory to recalculate")
            return

        # --- BACKUP originals ---
        originals = {r: copy.deepcopy(self.trajectories[r]) for r in rows}

        for r in rows:
            old = self.trajectories[r]
            centers_to_remove = [
                (f, cx, cy)
                for f, c in zip(old["frames"], old["search_centers"])
                if isinstance(c, (tuple,list)) and len(c)==2 and c[0] is not None
                for cx, cy in [c]
            ]
            self.navigator._remove_past_centers(centers_to_remove)

        # 2) single‐trajectory: do it inline on the GUI thread
        if len(rows) == 1:
            row = rows[0]
            old = self.trajectories[row]

            # — build the (frame,x,y) list as in worker _build_pts_for —
            anchors, roi = old["anchors"], old["roi"]
            nodes = old.get("nodes", [])
            if len(anchors) > 1 and roi is not None:
                pts = []
                for i in range(len(anchors) - 1):
                    f1, x1, _ = anchors[i]
                    f2, x2, _ = anchors[i+1]
                    seg = range(f1, f2+1) if i == 0 else range(f1+1, f2+1)
                    xs = np.linspace(x1, x2, len(seg), endpoint=True)
                    for j, f in enumerate(seg):
                        mx, my = compute_roi_point(roi, xs[j])
                        pts.append((f, mx, my))
            elif len(nodes) > 1:
                pts = [(f, x, y) for f, x, y in nodes]
            else:
                pts = [
                    (f, x, y)
                    for f, (x, y) in zip(old["frames"], old["original_coords"])
                ]

            try:
                trajectory_background = self.navigator.compute_trajectory_background(
                    self.navigator.get_movie_frame,
                    pts,
                    crop_size=int(2 * self.navigator.searchWindowSpin.value())
                )
                # — call compute_analysis on the GUI thread so its own dialog appears —
                frames, coords, centers, ints, fit, background = \
                    self.navigator._compute_analysis(pts, trajectory_background, showprogress=True)
            except Exception as e:
                print(f"Processing failed: {e}")
                self.navigator._is_canceled = True

            if self.navigator._is_canceled:
                self.navigator._is_canceled = False
                return

            # — rebuild the trajectory dict exactly as in _recalculate_one —
            spots   = [p[0] for p in fit]
            sigmas  = [p[1] for p in fit]
            peaks   = [p[2] for p in fit]
            valid   = [v for v,s in zip(ints, spots) if v and v>0 and s]
            avg_int = float(np.mean(valid)) if valid else None
            med_int = float(np.median(valid)) if valid else None

            vels = []
            for i in range(1, len(spots)):
                p0, p1 = spots[i-1], spots[i]
                if p0 is None or p1 is None:
                    vels.append(None)
                else:
                    vels.append(np.hypot(p1[0]-p0[0], p1[1]-p0[1]))
            good_vels = [v for v in vels if v is not None]
            avg_vpf   = float(np.mean(good_vels)) if good_vels else None

            traj_nodes = list(nodes)
            if len(anchors) > 1 and roi is not None:
                traj_nodes = []
                for frame, ax_x, _ax_y in anchors:
                    mx, my = compute_roi_point(roi, ax_x)
                    traj_nodes.append((int(frame), float(mx), float(my)))

            start = pts[0] if pts else old["start"]
            end = pts[-1] if pts else old["end"]

            traj_data = {
                "trajectory_number": old["trajectory_number"],
                "channel": old["channel"],
                "start":    start,
                "end":      end,
                "anchors":  anchors,
                "roi":      roi,
                "spot_centers": spots,
                "sigmas":      sigmas,
                "peaks":       peaks,
                "fixed_background": trajectory_background,
                "background": background,
                "frames":      frames,
                "original_coords": coords,
                "search_centers":  centers,
                "intensities":     ints,
                "average":         avg_int,
                "median":          med_int,
                "velocities":      vels,
                "average_velocity":avg_vpf,
                "nodes":            traj_nodes,
                "click_source":     old.get("click_source", "")
            }

            traj_data["custom_fields"] = originals[row].get("custom_fields", {}).copy()
            traj_data["segment_diffusion"] = originals[row].get("segment_diffusion", [])

            new_centers = [
                (f, cx, cy)
                for f, c in zip(traj_data["frames"], traj_data["search_centers"])
                if isinstance(c, (tuple,list)) and len(c)==2 and c[0] is not None
                for cx, cy in [c]
            ]
            self.navigator.past_centers.extend(new_centers)

            # After computing spots/ints/fit, before updateTableRow.
            if self.navigator.movie.ndim == 4 and self.navigator._channel_axis is not None:
                n_chan = self.navigator.movie.shape[self.navigator._channel_axis]
            else:
                n_chan = 1
            ref_ch = traj_data["channel"]

            # --- Custom colocalization block ---
            if getattr(self.navigator, "check_colocalization", False):
                # rehydrate analysis state
                self.navigator.analysis_frames     = traj_data["frames"]
                self.navigator.analysis_fit_params = list(zip(
                    traj_data["spot_centers"],
                    traj_data["sigmas"],
                    traj_data["peaks"]
                ))
                # compute everything in one pass exactly like run_analysis_points
                self.navigator.analysis_channel = ref_ch
                self.navigator._compute_colocalization(showprogress=True)
                any_list = list(self.navigator.analysis_colocalized)
                # grab the per-channel lists that _compute_colocalization just built
                by_ch = {
                    ch: list(flags)
                    for ch, flags in self.navigator.analysis_colocalized_by_ch.items()
                }
                traj_data["colocalization_any"]   = any_list
                traj_data["colocalization_by_ch"] = by_ch
                # compute per-trajectory percentages
                cf = traj_data.setdefault("custom_fields", {})
                valid_any = [s for s in any_list if s is not None]
                pct_any = f"{100*sum(1 for s in valid_any if s=='Yes')/len(valid_any):.1f}" if valid_any else ""
                for ch in range(1, n_chan+1):
                    key = f"Ch. {ch} co. %"
                    if ch == ref_ch:
                        cf[key] = ""
                    elif n_chan == 2:
                        cf[key] = pct_any
                    else:
                        flags = by_ch.get(ch, [])
                        valid = [s for s in flags if s is not None]
                        cf[key] = (f"{100*sum(1 for s in valid if s=='Yes')/len(valid):.1f}" if valid else "")
            else:
                any_list = [None] * len(traj_data["frames"])
                by_ch    = {
                    ch: [None] * len(traj_data["frames"])
                    for ch in range(1, n_chan+1) if ch != ref_ch
                }
                traj_data["colocalization_any"]   = any_list
                traj_data["colocalization_by_ch"] = by_ch

            # 4) Populate custom_fields for table/draw.
            # (already handled above in the colocalization block)

            if getattr(self.navigator, "show_steps", False):
                step_idxs, step_meds = self.navigator.compute_steps_for_data(
                    traj_data["frames"],
                    traj_data["intensities"]
                )
                traj_data["step_indices"] = step_idxs
                traj_data["step_medians"] = step_meds
            else:
                traj_data["step_indices"] = None
                traj_data["step_medians"] = None

            if getattr(self.navigator, "show_diffusion", False):
                if self.navigator.pixel_size is not None and self.navigator.frame_interval is not None:
                    traj_data["segment_diffusion"] = self._compute_segment_diffusion(
                        traj_data, self.navigator
                    )

            # — swap it in and refresh UI —
            self.updateTableRow(row, traj_data)
            if self.navigator.traj_overlay_button.isChecked():
                self.on_trajectory_selected_by_index(rows[0])
            return


        # 3) multi‐trajectory: optionally show RecalcDialog
        if prompt:
            mode = getattr(self.navigator, "tracking_mode", "Independent")
            rad  = self.navigator.searchWindowSpin.value()
            dlg  = RecalcDialog(mode, rad,
                                message=f"{len(rows)} trajectories need recalc",
                                parent=self)
            if dlg.exec_() != QDialog.Accepted:
                return
            # apply any mode/radius changes
            self.navigator.searchWindowSpin.setValue(dlg.new_radius)
            self.navigator.tracking_mode = dlg.new_mode
            if hasattr(self, "trackingModeCombo"):
                self.trackingModeCombo.setCurrentText(dlg.new_mode)

        # 4) set up master progress + thread
        total_frames = sum(len(self.trajectories[r]["frames"]) for r in rows)
        master = QProgressDialog("Recalculating", "Cancel", 0, total_frames, self)
        master.setWindowModality(Qt.WindowModal)
        master.setMinimumDuration(0)
        master.show()

        self.navigator._suppress_internal_progress = True

        worker = RecalcWorker(rows, self.trajectories, self.navigator)
        thread = QThread(self)
        worker.moveToThread(thread)
        self._recalc_thread = thread
        self._recalc_worker = worker

        worker.progress.connect(master.setValue)
        master.canceled.connect(worker.cancel)

        def on_finished(results):
            for row, traj_data in results:
                if self.navigator.movie.ndim == 4 and self.navigator._channel_axis is not None:
                    n_chan = self.navigator.movie.shape[self.navigator._channel_axis]
                else:
                    n_chan = 1
                ref_ch = traj_data["channel"]

                if getattr(self.navigator, "check_colocalization", False):
                    # rehydrate analysis state
                    self.navigator.analysis_frames     = traj_data["frames"]
                    self.navigator.analysis_fit_params = list(zip(
                        traj_data["spot_centers"],
                        traj_data["sigmas"],
                        traj_data["peaks"]
                    ))
                    self.navigator.analysis_channel    = ref_ch

                    # run colocalization in one pass
                    self.navigator._compute_colocalization(showprogress=False)
                    any_list = list(self.navigator.analysis_colocalized)
                    by_ch = {
                        ch: list(flags)
                        for ch, flags in self.navigator.analysis_colocalized_by_ch.items()
                    }
                else:
                    any_list = [None] * len(traj_data["frames"])
                    by_ch    = { ch: [None]*len(traj_data["frames"])
                                for ch in range(1, n_chan+1) if ch!=ref_ch }

                # store the raw flags
                traj_data["colocalization_any"]   = any_list
                traj_data["colocalization_by_ch"] = by_ch

                # *** NOW POPULATE the custom_fields columns ***
                cf = traj_data.setdefault("custom_fields", {})
                if n_chan == 2:
                    valid = [s for s in any_list if s is not None]
                    pct_any = f"{100*sum(1 for s in valid if s=='Yes')/len(valid):.1f}" if valid else ""
                    for ch in (1,2):
                        cf[f"Ch. {ch} co. %"] = "" if ch==ref_ch else pct_any
                else:
                    for ch in range(1, n_chan+1):
                        col = f"Ch. {ch} co. %"
                        if ch == ref_ch:
                            cf[col] = ""
                        else:
                            flags = by_ch.get(ch, [])
                            valid = [s for s in flags if s is not None]
                            cf[col] = (f"{100*sum(1 for s in valid if s=='Yes')/len(valid):.1f}" if valid else "")

                if getattr(self.navigator, "show_steps", False):
                    step_idxs, step_meds = self.navigator.compute_steps_for_data(
                        traj_data["frames"],
                        traj_data["intensities"]
                    )
                    traj_data["step_indices"] = step_idxs
                    traj_data["step_medians"] = step_meds
                else:
                    traj_data["step_indices"] = None
                    traj_data["step_medians"] = None

                # store back and refresh table
                self.trajectories[row] = traj_data
                self.updateTableRow(row, traj_data)

            cleanup()

        worker.finished.connect(on_finished)

        def on_canceled():
            # restore backups
            for row, old in originals.items():
                self.trajectories[row] = old
                self.updateTableRow(row, old)
            cleanup()

        worker.canceled.connect(on_canceled)

        def cleanup():
            master.close()
            self.navigator._suppress_internal_progress = False
            thread.quit()
            thread.wait()
            # redraw if overlay is on
            if self.navigator.traj_overlay_button.isChecked():
                self.on_trajectory_selected_by_index(rows[0])

        thread.started.connect(worker.run)
        thread.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        thread.start()

    def recalculate_all_trajectories(self):
        if not self.trajectories:
            QMessageBox.warning(self, "", "No Trajectories.")
            return

        # 1) Ask the user once for tracking mode / search radius
        mode = getattr(self.navigator, "tracking_mode", "Independent")
        rad  = self.navigator.searchWindowSpin.value()
        dlg  = RecalcDialog(mode, rad, message="", parent=self)
        if dlg.exec_() != QDialog.Accepted:
            return

        self.navigator.searchWindowSpin.setValue(dlg.new_radius)
        self.tracking_mode = dlg.new_mode
        if hasattr(self, "trackingModeCombo"):
            self.trackingModeCombo.setCurrentText(dlg.new_mode)

        # 2) Back up all current trajectories
        backup = copy.deepcopy(self.trajectories)

        # 3) Remove old “past_centers” from the navigator
        all_old = []
        for old in backup:
            all_old += [
                (f, cx, cy)
                for f, c in zip(old["frames"], old["search_centers"])
                if isinstance(c, (tuple, list)) and len(c) == 2 and c[0] is not None
                for cx, cy in [c]
            ]
        self.navigator._remove_past_centers(all_old)

        # 4) Create the dialog but do NOT show yet
        total_trajectories = len(backup)
        progress_dialog = QProgressDialog("Recalculating all…", "Cancel", 0, total_trajectories, self)
        progress_dialog.setCancelButton(None) #IT DOESNT QUIT OUT OF WHOLE PROGRESS, HAD TO REMOVE
        progress_dialog.setWindowModality(Qt.WindowModal)
        progress_dialog.setMinimumDuration(0)

        # 5) Create exactly one worker and thread, and store them on self
        self._recalc_backup = backup
        self._recalc_worker = RecalcAllWorker(self._recalc_backup, self.navigator)
        self._recalc_thread = QThread(self)
        self._recalc_worker.moveToThread(self._recalc_thread)

        # 6) Define callbacks that refer to self._recalc_thread
        def on_worker_finished(results: dict):
            # Swap new trajectories back into the model & update the table UI
            for row_idx, new_traj in results.items():
                self.trajectories[row_idx] = new_traj
                self.updateTableRow(row_idx, new_traj)

            progress_dialog.close()
            self.navigator._suppress_internal_progress = False

            # Now stop the thread we actually used:
            self._recalc_thread.quit()
            self._recalc_thread.wait()

            # Re-select row 0.
            self.on_trajectory_selected_by_index(0)
            self.table_widget.selectRow(0)

        def on_worker_canceled():
            # Restore backup if the user hit Cancel
            for i, orig in enumerate(self._recalc_backup):
                self.trajectories[i] = orig
                self.updateTableRow(i, orig)

            progress_dialog.close()
            self.navigator._suppress_internal_progress = False

            # Also quit the same thread:
            self._recalc_thread.quit()
            self._recalc_thread.wait()

        # 7) Hook up all signals (including dialog.canceled → worker.cancel)
        self._recalc_worker.progress.connect(progress_dialog.setValue)
        self._recalc_worker.finished.connect(on_worker_finished)
        self._recalc_worker.canceled.connect(on_worker_canceled)
        progress_dialog.canceled.connect(self._recalc_worker.cancel)

        # 8) Show the dialog, then start the thread running our worker
        progress_dialog.show()
        self.navigator._suppress_internal_progress = True

        self._recalc_thread.started.connect(self._recalc_worker.run)
        self._recalc_thread.finished.connect(self._recalc_worker.deleteLater)
        self._recalc_thread.finished.connect(self._recalc_thread.deleteLater)
        self._recalc_thread.start()

    def toggle_trajectory_markers(self):
        mode = self.navigator.get_traj_overlay_mode() if self.navigator is not None else "all"
        if mode == "off":
            self.kymoCanvas.clear_kymo_trajectory_markers()
            self.movieCanvas.clear_movie_trajectory_markers()
        else:
            self.kymoCanvas.draw_trajectories_on_kymo()
            if self.navigator is not None:
                self.movieCanvas.draw_trajectories_on_movie()
        
        self.movieCanvas.draw()
        self.kymoCanvas.draw()
        if self.navigator is not None:
            try:
                self.navigator._rebuild_movie_blit_background()
            except Exception:
                pass

    def delete_selected_trajectory(self):
        # 1) Get all selected rows
        selected_rows = [idx.row() for idx in self.table_widget.selectionModel().selectedRows()]
        if not selected_rows:
            # nothing selected
            return

        # 2) Sort in reverse so removing rows doesn't shift indices of earlier ones
        selected_rows.sort(reverse=True)

        # 3) Prepare to collect all spot-centers to remove
        centers_to_remove = []

        # 4) Remove each selected trajectory
        for row in selected_rows:
            deleted = self.trajectories.pop(row)
            # remove from table
            self.table_widget.removeRow(row)
            # gather valid (x,y) centers
            for frame, center in zip(deleted["frames"], deleted["spot_centers"]):
                if isinstance(center, (tuple, list)) and len(center) == 2 and center[0] is not None and center[1] is not None:
                    centers_to_remove.append((frame, center[0], center[1]))

        # 5) Tell navigator to drop all those centers
        if centers_to_remove and self.navigator is not None:
            self.navigator._remove_past_centers(centers_to_remove)

        # 6) Re-select a sensible row in the table
        row_count = self.table_widget.rowCount()
        if row_count > 0:
            # pick the smallest index of those we deleted, clamped to [0, row_count-1]
            new_row = min(selected_rows)
            new_row = max(0, min(new_row, row_count - 1))
            self.table_widget.selectRow(new_row)
            self.current_index = new_row
        else:
            self.table_widget.clearSelection()
            self.current_index = None

        # 7) Recompute trajectory counter:
        #    if none left, reset to 1; otherwise max+1
        if not self.trajectories:
            self._trajectory_counter = 1
        else:
            max_num = max(t["trajectory_number"] for t in self.trajectories)
            self._trajectory_counter = max_num + 1

        # 8) Redraw everything
        self.movieCanvas.draw_trajectories_on_movie()
        if self.navigator is not None:
            self.kymoCanvas.draw_trajectories_on_kymo()
            self.navigator.update_table_visibility()

        self.movieCanvas.draw()
        self.kymoCanvas.draw()

    def clear_trajectories(self, prompt=True):
        reply = QMessageBox.Yes
        if prompt:
            reply = QMessageBox.question(
                self,
                "Delete Trajectories",
                "Are you sure you want to delete all trajectories?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
        if reply == QMessageBox.Yes or not prompt:

            for col_name in ("Number of Steps", "Average Step Size"):
                if col_name in self.custom_columns:
                    idx = self._col_index[col_name]
                    # remove from the Qt table
                    self.table_widget.removeColumn(idx)
                    # drop from our header & index map
                    self._headers.pop(idx)
                    self._column_types.pop(col_name, None)
                    self.custom_columns.remove(col_name)
                    # rebuild the lookup
                    self._col_index = {hdr: i for i, hdr in enumerate(self._headers)}

            for traj in self.trajectories:
                for key in list(traj.keys()):
                    if key == "custom_fields":
                        for col in self.custom_columns:
                            traj["custom_fields"][col] = ""
                    else:
                        traj[key] = None
            self.trajectories = []
            self.table_widget.setRowCount(0)
            self._trajectory_counter = 1
            if self.navigator is not None:
                self.kymoCanvas.clear_kymo_trajectory_markers()
                self.movieCanvas.clear_movie_trajectory_markers()
            self.navigator.update_table_visibility()

            self.kymoCanvas.draw()
            self.movieCanvas.draw()

            self.navigator.past_centers = []

    def update_trajectory_visibility(self):
        has_rows = self.table_widget.rowCount() > 0
        if not has_rows:
            # Hide (collapse) the trajectory canvas by setting its size to zero
            total_height = self.navigator.vertSplitter.height()
            self.navigator.vertSplitter.setSizes([total_height, 0])
        else:
            # Give some proportion (e.g., 70% to main content, 30% to trajectory canvas)
            total_height = self.navigator.vertSplitter.height()
            self.navigator.vertSplitter.setSizes([int(0.7 * total_height), int(0.3 * total_height)])

    def open_context_menu(self, pos):
        # 1) Figure out which row was clicked on
        index_under = self.table_widget.indexAt(pos)
        if not index_under.isValid():
            return
        row_under = index_under.row()

        # 2) Determine the “rows” list
        selected_indexes = self.table_widget.selectionModel().selectedRows()
        if any(idx.row() == row_under for idx in selected_indexes):
            rows = [idx.row() for idx in selected_indexes]
        else:
            rows = [row_under]

        # 3) Build the menu
        menu  = QMenu(self.table_widget)
        menu.setWindowFlags(menu.windowFlags() | Qt.FramelessWindowHint)
        menu.setAttribute(Qt.WA_TranslucentBackground)
        menu.setAttribute(Qt.WA_NoMouseReplay)

        def _add_action(label, callback):
            act = menu.addAction(label)
            act.triggered.connect(lambda _chk=False, cb=callback: cb())
            return act

        # --- (A) Save action ---
        n = len(rows)
        save_label = (f"Save trajectory {rows[0]+1}" 
                    if n == 1 
                    else "Save selected trajectories")
        _add_action(
            save_label,
            lambda sel=rows: self.save_trajectories(sel),
        )

        # --- (B) “Check colocalization” for single row if needed ---
        if n == 1 and getattr(self.navigator, "check_colocalization", False) and self.navigator.movie.ndim == 4:
            r    = rows[0]
            traj = self.trajectories[r]
            cf   = traj.get("custom_fields", {})
            ref_ch = traj["channel"]

            # test for any missing co. % (skip the ref channel col)
            missing = any(
                col.endswith(" co. %")
                and not col.endswith(f"{ref_ch} co. %")
                and not cf.get(col, "").strip()
                for col in self.custom_columns
            )
            if missing:
                def _do_coloc_check(row):
                    self.navigator._compute_colocalization_for_row(row)
                    self.navigator.kymoCanvas.draw_trajectories_on_kymo()
                    self.navigator.kymoCanvas.draw_idle()
                _add_action(
                    "Check colocalization",
                    lambda row=r: _do_coloc_check(row),
                )
                menu.addSeparator()

        # --- (C) Go→kymograph entries if exactly one row ---
        if n == 1:
            r    = rows[0]
            traj = self.trajectories[r]
            sf, sx, sy = traj["start"]
            ef, ex, ey = traj["end"]
            traj_ch     = traj["channel"]

            for roi_name, roi in self.navigator.rois.items():
                if is_point_near_roi((sx, sy), roi) and is_point_near_roi((ex, ey), roi):
                    for kymo_name, info in self.navigator.kymo_roi_map.items():
                        if info["roi"] == roi_name and info.get("channel") == traj_ch:
                            def _go_to_kymo(kn, row):
                                self.navigator.kymoCombo.setCurrentIndex(
                                    self.navigator.kymoCombo.findText(kn)
                                )
                                # kymo_changed is connected to the combo; let it run first
                                self.on_trajectory_selected_by_index(row)
                            _add_action(
                                f"Go to kymograph {kymo_name}",
                                lambda kn=kymo_name, row=r: _go_to_kymo(kn, row),
                            )
                    break  # only first matching ROI

        # --- (D) custom‐column toggles / setters as before ---
        if self.custom_columns:
            menu.addSeparator()
            seen = set()
            # Don’t offer manual setters for computed diffusion columns
            d_col = getattr(self.navigator, "_DIFF_D_COL", "Diffusion D (µm²/s)")
            a_col = getattr(self.navigator, "_DIFF_A_COL", "Diffusion α")
            _computed_cols = {d_col, a_col}
            for col_name in self.custom_columns:
                if col_name in _computed_cols:
                    continue
                col_type = self._column_types.get(col_name, "binary")
                if col_type == "binary":
                    marked_flags = [
                        bool(self.trajectories[r]["custom_fields"].get(col_name, ""))
                        for r in rows
                    ]
                    if n == 1:
                        act_text = (f"Unmark as {col_name}" if marked_flags[0]
                                    else f"Mark as {col_name}")
                    else:
                        act_text = (f"Unmark selected as {col_name}"
                                    if all(marked_flags) 
                                    else f"Mark selected as {col_name}")

                    if act_text not in seen:
                        seen.add(act_text)
                        _add_action(
                            act_text,
                            lambda name=col_name, sel=rows: self._toggle_binary_column(name, sel),
                        )

                elif col_type == "value":
                    if n == 1:
                        act_text = f"Set {col_name}"
                    else:
                        act_text = f"Set all {col_name}"
                    if act_text not in seen:
                        seen.add(act_text)
                        _add_action(
                            act_text,
                            lambda name=col_name, sel=rows: self._set_value_column(name, sel),
                        )

        # 4) Pop up
        menu.exec_(self.table_widget.viewport().mapToGlobal(pos))

    def _toggle_binary_column(self, col_name, rows):
        """
        For each row in `rows`, either mark or unmark the binary column `col_name`.
        """
        # decide whether to mark (True) or unmark (False)
        current = [
            bool(self.trajectories[r]["custom_fields"].get(col_name, ""))
            for r in rows
        ]
        mark = not all(current)

        for r in rows:
            if mark:
                # Set to "Yes" or a non-empty marker.
                self._mark_custom(r, col_name, "Yes")
            else:
                self._unmark_custom(r, col_name)

        if self.navigator.color_by_column == col_name:
            self.navigator.refresh_color_by()
            return

        self.kymoCanvas.draw_trajectories_on_kymo()
        self.kymoCanvas.draw()
        self.movieCanvas.draw_trajectories_on_movie()
        self.movieCanvas.draw()
        self.navigator._update_legends()
        
    def _set_value_column(self, col_name, rows):
        """
        Pop up a styled dialog to get a new value, then set it on every row in `rows`.
        """
        # Prevent manual editing of computed diffusion columns
        d_col = getattr(self.navigator, "_DIFF_D_COL", "Diffusion D (µm²/s)")
        a_col = getattr(self.navigator, "_DIFF_A_COL", "Diffusion α")
        if col_name in (d_col, a_col):
            QMessageBox.information(self, "", f"{col_name} is computed and cannot be set manually.")
            return

        prompt = f"Enter value for {col_name}:"
        # 1) Create QInputDialog instance
        dlg = QInputDialog(self)
        dlg.setWindowTitle(f"Set {col_name}")
        dlg.setLabelText(prompt)
        dlg.setOkButtonText("OK")
        dlg.setCancelButtonText("Cancel")
        dlg.setInputMode(QInputDialog.TextInput)
        dlg.setTextValue("")  # start empty

        # 2) Style its line‐edit to have a white background
        line: QLineEdit = dlg.findChild(QLineEdit)
        if line:
            line.setStyleSheet("background-color: white;")

        # 3) Show and handle the result
        if dlg.exec_() != QDialog.Accepted:
            return
        val = dlg.textValue()

        # 4) Store and update
        for r in rows:
            self.trajectories[r].setdefault("custom_fields", {})[col_name] = val
            self.writeToTable(r, col_name, val)

        if self.navigator.color_by_column == col_name:
            self.navigator.refresh_color_by()
            return

        self.kymoCanvas.draw_trajectories_on_kymo()
        self.kymoCanvas.draw()
        self.movieCanvas.draw_trajectories_on_movie()
        self.movieCanvas.draw()
        if self.navigator.color_by_column == col_name:
            self.navigator._update_legends()

    def _on_header_context_menu(self, pos):
        header = self.table_widget.horizontalHeader()
        col = header.logicalIndexAt(pos)
        if col < 0:
            return

        col_name = self._headers[col]
        menu = QMenu(self.table_widget)
        # make it frameless + translucent
        menu.setWindowFlags(menu.windowFlags() | Qt.FramelessWindowHint)
        menu.setAttribute(Qt.WA_TranslucentBackground)
        menu.setAttribute(Qt.WA_NoMouseReplay)

        # always allow adding a new column
        # 1) Add binary column
        bin_act = menu.addAction("Add Binary column")
        bin_act.triggered.connect(self._add_binary_column_dialog)

        # 2) Add value column
        val_act = menu.addAction("Add Value column")
        val_act.triggered.connect(self._add_value_column_dialog)

        # if this is one of the custom columns, allow removal
        ctype = self._column_types.get(col_name)
        # only offer “remove” for our binary/value custom columns
        if col_name in self.custom_columns and ctype in ("binary", "value"):
            remove_act = menu.addAction(f"Remove column {col_name}")
            remove_act.triggered.connect(lambda _, c=col, name=col_name: self._ask_remove_column(c, name))

        # finally show the menu right at the header click
        menu.exec_(header.mapToGlobal(pos))

    def _add_binary_column_dialog(self):
        # 1) Create QInputDialog instance
        dlg = QInputDialog(self)
        dlg.setWindowTitle("Add Binary Column")
        dlg.setLabelText("Name:")
        dlg.setOkButtonText("OK")
        dlg.setCancelButtonText("Cancel")
        dlg.setInputMode(QInputDialog.TextInput)
        dlg.setTextValue("")

        # 2) style its line‐edit to have a white background
        line = dlg.findChild(QLineEdit)
        if line:
            line.setStyleSheet("background-color: white;")

        # 3) exec and check result
        if dlg.exec_() != QDialog.Accepted:
            return
        name = dlg.textValue().strip()
        if not name:
            return

        # 4) duplicate‐check
        if name.lower() in {h.lower() for h in self._headers}:
            QMessageBox.warning(self, "Duplicate", f"“{name}” already exists.")
            return

        # 5) finally, add the binary column
        self._add_custom_column(name, col_type="binary")

    def _add_value_column_dialog(self):
        # 1) Create QInputDialog instance
        dlg = QInputDialog(self)
        dlg.setWindowTitle("Add Value Column")
        dlg.setLabelText("Name:")
        dlg.setOkButtonText("OK")
        dlg.setCancelButtonText("Cancel")
        dlg.setInputMode(QInputDialog.TextInput)
        dlg.setTextValue("")          # start empty

        # 2) style its line‐edit to have a white background
        line = dlg.findChild(QLineEdit)
        if line:
            line.setStyleSheet("background-color: white;")

        # 3) exec and check result
        if dlg.exec_() != QDialog.Accepted:
            return
        name = dlg.textValue().strip()
        if not name:
            return

        # 4) duplicate‐check
        if name.lower() in {h.lower() for h in self._headers}:
            QMessageBox.warning(self, "Duplicate", f"“{name}” already exists.")
            return

        # 5) add the new column
        self._add_custom_column(name, col_type="value")

    # def _add_column_dialog(self):
    #     """
    #     Ask for a new column name, but only accept it if it doesn't
    #     case-insensitively collide with any existing header.
    #     """
    #     existing = {h.lower() for h in self._headers}
    #     name, ok = QInputDialog.getText(self, "Add Column", "Column name:")
    #     if not ok:
    #         return
    #     name = name.strip()
    #     if not name:
    #         return
    #     if name.lower() in existing:
    #         QMessageBox.warning(self, "Duplicate column", f"'{name}' already exists.")
    #         return

    #     self._add_custom_column(name)

    def _add_custom_column(self, name, *, col_type="binary"):

        # print("_add_custom_column", "name, col_type", name, col_type)
        
        # ——— 0) save current row selection ———
        selected_rows = [idx.row()
                         for idx in self.table_widget.selectionModel().selectedRows()]
        
        self._column_types[name] = col_type

        # 1) track it
        self.custom_columns.append(name)

        # print("_add_custom_column", "self.custom_columns", self.custom_columns)

        # 2) extend header lists & mappings
        self._headers.append(name)
        idx = len(self._headers) - 1
        self._col_index[name] = idx
        self._aliases[name.lower()] = name

        # 3) insert the column into the table widget
        self.table_widget.insertColumn(idx)
        self.table_widget.setHorizontalHeaderLabels(self._headers)

        # 4) give it a reasonable default width & make it resizable
        self.table_widget.horizontalHeader().setSectionResizeMode(
            idx, QHeaderView.Interactive)
        self.table_widget.setColumnWidth(idx, 80)

        # 5) initialize existing rows to empty string *and* store in data model
        default = "" if col_type=="binary" else ""
        for row in range(self.table_widget.rowCount()):
            self.writeToTable(row, name, default)
            self.trajectories[row].setdefault("custom_fields", {})[name] = default

        # 6) re-apply full-row selection so that the new column is highlighted
        sel = self.table_widget.selectionModel()
        sel.blockSignals(True)
        for row in selected_rows:
            self.table_widget.selectRow(row)
        sel.blockSignals(False)

        # 7) final layout tweaks
        self.table_widget.viewport().update()
        # self.table_widget.horizontalHeader().resizeSections(
        #     QHeaderView.ResizeToContents)

        self.navigator._rebuild_color_by_actions()

    def _ask_remove_column(self, col_idx, col_name):
        # Check for any non-empty cells
        has_values = any(
            (self.table_widget.item(r, col_idx) or QTableWidgetItem("")).text().strip()
            for r in range(self.table_widget.rowCount())
        )
        if has_values:
            reply = QMessageBox.question(
                self,
                f"Remove “{col_name}”?",
                f"Column “{col_name}” still has values. Remove and lose all data?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                return

        self._remove_custom_column(col_idx, col_name)

    def _remove_custom_column(self, col_idx, col_name):
        # 1) remove from data model
        self.custom_columns.remove(col_name)
        self._headers.pop(col_idx)
        self._col_index.pop(col_name, None)
        self._aliases.pop(col_name.lower(), None)

        # 2) drop from each trajectory
        for traj in self.trajectories:
            traj.get("custom_fields", {}).pop(col_name, None)

        # 3) remove from widget and rebuild our index map
        self.table_widget.removeColumn(col_idx)
        # rebuild header→index map so no one’s left pointing at the wrong slot
        self._col_index = { hdr: i for i, hdr in enumerate(self._headers) }
        self.table_widget.setHorizontalHeaderLabels(self._headers)
        self.table_widget.viewport().update()

        self.navigator._rebuild_color_by_actions()

    def _mark_custom(self, row, col_name, value="Yes"):
        # 1) update data model
        self.trajectories[row].setdefault("custom_fields", {})[col_name] = value
        # 2) update UI
        self.writeToTable(row, col_name, value)

    def _unmark_custom(self, row, column_name):
        # 1) update data-model: clear or remove the custom_fields entry
        cf = self.trajectories[row].get("custom_fields", {})
        if column_name in cf:
            # either delete it entirely...
            cf.pop(column_name)
            # ...or set to empty string to keep the key:
            # cf[column_name] = ""
        # 2) update the widget
        self.writeToTable(row, column_name, "")
